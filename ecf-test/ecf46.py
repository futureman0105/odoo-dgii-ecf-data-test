import openpyxl
import xml.etree.ElementTree as ET
import logging
import os
from xml.dom import minidom
from utility import clean_xml_safe 

workbook = openpyxl.load_workbook("test_data.xlsx")
sheet = workbook["ECF"]

class ECF46:
    """Class to handle E-CF-34 XML generation."""
    def __init__(self):
        self.RNCEmisor = ""
        self.ENCF = ""
        self.RNCComprador = ""
        self.invoice_name = ""
        
        
    def create_e_cf_46(self, in_row : int) :

        """Generate DGII-compliant XML from an Odoo invoice."""
        # Create root element with namespaces
        root = ET.Element('ECF', {
            'xmlns:xs' : "http://www.w3.org/2001/XMLSchema",
        })

        # 1. Encabezado
        encabezado = ET.SubElement(root, 'Encabezado')

        # Write Version
        search_text = "Version" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value != "#e" :    
                    ET.SubElement(encabezado, 'Version').text = value
                    print(f'Version : {value}')
                    print(f'Version Col Index : {col}')
                break

        # IdDoc
        id_doc = ET.SubElement(encabezado, 'IdDoc')

        # Write TipoeCF
        search_text = "TipoeCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 
                ET.SubElement(id_doc, 'TipoeCF').text = value
                print(f'TipoeCF : {cell_value}')
                break

        # Write eNCF
        search_text = "ENCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 
                ET.SubElement(id_doc, 'eNCF').text = value
                self.ENCF = value
                print(f'eNCF : {cell_value}')
                break

        # Write FechaVencimientoSecuencia
        search_text = "FechaVencimientoSecuencia" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'FechaVencimientoSecuencia').text = value
                print(f'FechaVencimientoSecuencia : {value}')
                break

        # Write IndicadorEnvioDiferido
        search_text = "IndicadorEnvioDiferido" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'IndicadorEnvioDiferido').text = value
                print(f'IndicadorEnvioDiferido : {value}')
                break

        # Write TipoIngresos
        search_text = "TipoIngresos" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TipoIngresos').text = value
                print(f'TipoIngresos : {value}')
                break

        # Write TipoPago
        search_text = "TipoPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TipoPago').text = value
                print(f'TipoPago : {value}')
                break

        # Write FechaLimitePago
        search_text = "FechaLimitePago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'FechaLimitePago').text = value
                print(f'FechaLimitePago : {value}')
                break

        # Write TerminoPago
        search_text = "TerminoPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TerminoPago').text = value
                print(f'TerminoPago : {value}')
                break
        
        
        TablaFormasPago = ET.SubElement(id_doc, 'TablaFormasPago')

        search_text = "FormaPago[1]"
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                FormaDePago_count = 7
                col_index = col
                while True :
                    FormaDePago_count -= 1
                    if FormaDePago_count < 0 :
                        break            
                    FormaDePago = ET.SubElement(TablaFormasPago, 'FormaDePago')

                    value = str(sheet.cell(in_row, column= col_index ).value)
                    if value == "#e" :
                        value = ""
                    ET.SubElement(FormaDePago, 'FormaPago').text = value
                    col_index += 1

                    value = str(sheet.cell(in_row, column= col_index).value)
                    if value == "#e" :
                        value = ""
                    ET.SubElement(FormaDePago, 'MontoPago').text = value
                    col_index += 1
        
        # Write TipoCuentaPago
        search_text = "TipoCuentaPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TipoCuentaPago').text = value
                print(f'TipoCuentaPago : {value}')
                break
            
        # Write NumeroCuentaPago
        search_text = "NumeroCuentaPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'NumeroCuentaPago').text = value
                print(f'NumeroCuentaPago : {value}')
                break
            
        # Write BancoPago
        search_text = "BancoPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'BancoPago').text = value
                print(f'BancoPago : {value}')
                break

        # Write FechaDesde
        search_text = "FechaDesde" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'FechaDesde').text = value
                print(f'FechaDesde : {value}')
                break    

        # Write FechaHasta
        search_text = "FechaHasta" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'FechaHasta').text = value
                print(f'FechaHasta : {value}')
                break    

        # Write TotalPaginas
        search_text = "TotalPaginas" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TotalPaginas').text = value
                print(f'TotalPaginas : {value}')
                break

        Emisor = ET.SubElement(encabezado, 'Emisor')

        # Write RNCEmisor
        search_text = "RNCEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RNCEmisor').text = value
                self.RNCEmisor = value
                print(f'RNCEmisor : {value}')
                break

        # Write RazonSocialEmisor
        search_text = "RazonSocialEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RazonSocialEmisor').text = value
                print(f'RazonSocialEmisor : {value}')
                break

        # Write NombreComercial
        search_text = "NombreComercial" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NombreComercial').text = value
                print(f'NombreComercial : {value}')
                break

        # Write Sucursal
        search_text = "Sucursal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Sucursal').text = value
                print(f'Sucursal : {value}')
                break

        # Write DireccionEmisor
        search_text = "DireccionEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'DireccionEmisor').text = value
                print(f'DireccionEmisor : {value}')
                break

        # Write Municipio
        search_text = "Municipio" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Municipio').text = value
                print(f'Municipio : {value}')
                break

        # Write Provincia
        search_text = "Provincia" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Provincia').text = value
                print(f'Provincia : {value}')
                break

        TablaTelefonoEmisor = ET.SubElement(Emisor, 'TablaTelefonoEmisor')

        # Write TelefonoEmisor
        search_text = "TelefonoEmisor[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                TelefonoEmisor_count = 3
                col_index = col
                while True :
                    TelefonoEmisor_count -= 1
                    if TelefonoEmisor_count < 0 : 
                        break

                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""
                    ET.SubElement(TablaTelefonoEmisor, 'TelefonoEmisor').text = value           
                    print(f'TelefonoEmisor : {value}')

                    col_index +=1
                
        # Write CorreoEmisor
        search_text = "CorreoEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'CorreoEmisor').text = value
                print(f'CorreoEmisor : {value}')
                break

        # Write WebSite
        search_text = "WebSite" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'WebSite').text = value
                print(f'WebSite : {value}')
                break

        # Write ActividadEconomica
        search_text = "ActividadEconomica" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'ActividadEconomica').text = value
                print(f'ActividadEconomica : {value}')
                break

        # Write CodigoVendedor
        search_text = "CodigoVendedor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'CodigoVendedor').text = value
                print(f'CodigoVendedor : {value}')
                break

        # Write NumeroFacturaInterna
        search_text = "NumeroFacturaInterna" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NumeroFacturaInterna').text = value
                print(f'NumeroFacturaInterna : {value}')
                break
    
        # Write NumeroPedidoInterno
        search_text = "NumeroPedidoInterno" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NumeroPedidoInterno').text = value
                print(f'NumeroPedidoInterno : {value}')
                break
    
        # Write ZonaVenta
        search_text = "ZonaVenta" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'ZonaVenta').text = value
                print(f'ZonaVenta : {value}')
                break

        # Write RutaVenta
        search_text = "RutaVenta" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RutaVenta').text = value
                print(f'RutaVenta : {value}')
                break

        # Write InformacionAdicionalEmisor
        search_text = "InformacionAdicionalEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'InformacionAdicionalEmisor').text = value
                print(f'InformacionAdicionalEmisor : {value}')
                break

        # Write FechaEmision
        search_text = "FechaEmision" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'FechaEmision').text = value
                print(f'FechaEmision : {value}')
                break

        Comprador = ET.SubElement(encabezado, 'Comprador')

        # Write RNCComprador
        search_text = "RNCComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'RNCComprador').text = value
                self.RNCComprador = value
                print(f'RNCComprador : {value}')
                break

        # Write IdentificadorExtranjero
        search_text = "IdentificadorExtranjero" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'IdentificadorExtranjero').text = value
                print(f'IdentificadorExtranjero : {value}')
                break
    
        # Write RazonSocialComprador
        search_text = "RazonSocialComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'RazonSocialComprador').text = value
                print(f'RazonSocialComprador : {value}')
                break
    
        # Write ContactoComprador
        search_text = "ContactoComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'ContactoComprador').text = value
                print(f'ContactoComprador : {value}')
                break

        # Write CorreoComprador
        search_text = "CorreoComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'CorreoComprador').text = value
                print(f'CorreoComprador : {value}')
                break

        # Write DireccionComprador
        search_text = "DireccionComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'DireccionComprador').text = value
                print(f'DireccionComprador : {value}')
                break

        # Write MunicipioComprador
        search_text = "MunicipioComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'MunicipioComprador').text = value
                print(f'MunicipioComprador : {value}')
                break

        # Write ProvinciaComprador
        search_text = "ProvinciaComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'ProvinciaComprador').text = value
                print(f'ProvinciaComprador : {value}')
                break

        # Write FechaEntrega
        search_text = "FechaEntrega" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'FechaEntrega').text = value
                print(f'FechaEntrega : {value}')
                break

        # Write ContactoEntrega
        search_text = "ContactoEntrega" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'ContactoEntrega').text = value
                print(f'ContactoEntrega : {value}')
                break

        # Write DireccionEntrega
        search_text = "DireccionEntrega" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'DireccionEntrega').text = value
                print(f'DireccionEntrega : {value}')
                break

        # Write TelefonoAdicional
        search_text = "TelefonoAdicional" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'TelefonoAdicional').text = value
                print(f'TelefonoAdicional : {value}')
                break

        # Write FechaOrdenCompra
        search_text = "FechaOrdenCompra" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'FechaOrdenCompra').text = value
                print(f'FechaOrdenCompra : {value}')
                break

        # Write NumeroOrdenCompra
        search_text = "NumeroOrdenCompra" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'NumeroOrdenCompra').text = value
                print(f'NumeroOrdenCompra : {value}')
                break

        # Write CodigoInternoComprador
        search_text = "CodigoInternoComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'CodigoInternoComprador').text = value
                print(f'CodigoInternoComprador : {value}')
                break

        # Write ResponsablePago
        search_text = "ResponsablePago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'ResponsablePago').text = value
                print(f'ResponsablePago : {value}')
                break
        
        # Write InformacionAdicionalComprador
        search_text = "InformacionAdicionalComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'InformacionAdicionalComprador').text = value
                print(f'InformacionAdicionalComprador : {value}')
                break
        
        InformacionesAdicionales = ET.SubElement(encabezado, 'InformacionesAdicionales')
    
        # Write FechaEmbarque
        search_text = "FechaEmbarque" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'FechaEmbarque').text = value
                print(f'FechaEmbarque : {value}')
                break

        # Write NumeroEmbarque
        search_text = "NumeroEmbarque" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'NumeroEmbarque').text = value
                print(f'NumeroEmbarque : {value}')
                break

        # Write NumeroContenedor
        search_text = "NumeroContenedor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'NumeroContenedor').text = value
                print(f'NumeroContenedor : {value}')
                break

        value = str(sheet.cell(in_row, column=73).value)
        if value == "#e" :
            value = "" 

        ET.SubElement(InformacionesAdicionales, 'NumeroContenedor').text = value
        print(f'NumeroContenedor : {value}')
        
        # Write NumeroReferencia
        search_text = "NumeroReferencia" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'NumeroReferencia').text = value
                print(f'NumeroReferencia : {value}')
                break

        # Write PesoBruto
        search_text = "PesoBruto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'PesoBruto').text = value
                print(f'PesoBruto : {value}')
                break

        # Write PesoNeto
        search_text = "PesoNeto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'PesoNeto').text = value
                print(f'PesoNeto : {value}')
                break

        # Write UnidadPesoBruto
        search_text = "UnidadPesoBruto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'UnidadPesoBruto').text = value
                print(f'UnidadPesoBruto : {value}')
                break

        # Write UnidadPesoNeto
        search_text = "UnidadPesoNeto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'UnidadPesoNeto').text = value
                print(f'UnidadPesoNeto : {value}')
                break

        # Write CantidadBulto
        search_text = "CantidadBulto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'CantidadBulto').text = value
                print(f'CantidadBulto : {value}')
                break

        # Write UnidadBulto
        search_text = "UnidadBulto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'UnidadBulto').text = value
                print(f'UnidadBulto : {value}')
                break

        # Write VolumenBulto
        search_text = "VolumenBulto" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'VolumenBulto').text = value
                print(f'VolumenBulto : {value}')
                break

        # Write UnidadVolumen
        search_text = "UnidadVolumen" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(InformacionesAdicionales, 'UnidadVolumen').text = value
                print(f'UnidadVolumen : {value}')
                break

        # Transporte
        Transporte = ET.SubElement(encabezado, 'Transporte')

        # Write Conductor
        search_text = "Conductor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'Conductor').text = value
                print(f'Conductor : {value}')
                break

        # Write DocumentoTransporte
        search_text = "DocumentoTransporte" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'DocumentoTransporte').text = value
                print(f'DocumentoTransporte : {value}')
                break

        # Write Ficha
        search_text = "Ficha" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'Ficha').text = value
                print(f'Ficha : {value}')
                break

        # Write Placa
        search_text = "Placa" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'Placa').text = value
                print(f'Placa : {value}')
                break

        # Write RutaTransporte
        search_text = "RutaTransporte" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'RutaTransporte').text = value
                print(f'RutaTransporte : {value}')
                break

        # Write ZonaTransporte
        search_text = "ZonaTransporte" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'ZonaTransporte').text = value
                print(f'ZonaTransporte : {value}')
                break

        # Write NumeroAlbaran
        search_text = "NumeroAlbaran" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Transporte, 'NumeroAlbaran').text = value
                print(f'NumeroAlbaran : {value}')
                break

        Totales = ET.SubElement(encabezado, 'Totales')
    
        # Write MontoGravadoTotal
        search_text = "MontoGravadoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoGravadoTotal').text = value
                print(f'MontoGravadoTotal : {value}')
                break

        # Write MontoGravadoI3
        search_text = "MontoGravadoI3" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoGravadoI3').text = value
                print(f'MontoGravadoI3 : {value}')
                break

        # Write ITBIS3
        search_text = "ITBIS3" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'ITBIS3').text = value
                print(f'ITBIS3 : {value}')
                break
    
        # Write TotalITBIS
        search_text = "TotalITBIS" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'TotalITBIS').text = value
                print(f'TotalITBIS : {value}')
                break

        # Write TotalITBIS3
        search_text = "TotalITBIS3" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'TotalITBIS3').text = value
                print(f'TotalITBIS3 : {value}')
                break

        # Write MontoTotal
        search_text = "MontoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoTotal').text = value
                print(f'MontoTotal : {value}')
                break

        # Write MontoNoFacturable
        search_text = "MontoNoFacturable" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoNoFacturable').text = value
                print(f'MontoNoFacturable : {value}')
                break

        # Write MontoPeriodo
        search_text = "MontoPeriodo" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoPeriodo').text = value
                print(f'MontoPeriodo : {value}')
                break

        # Write SaldoAnterior
        search_text = "SaldoAnterior" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'SaldoAnterior').text = value
                print(f'SaldoAnterior : {value}')
                break

        # Write MontoAvancePago
        search_text = "MontoAvancePago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoAvancePago').text = value
                print(f'MontoAvancePago : {value}')
                break

        # Write ValorPagar
        search_text = "ValorPagar" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'ValorPagar').text = value
                print(f'ValorPagar : {value}')
                break

        OtraMoneda = ET.SubElement(encabezado, 'OtraMoneda')
    
        # Write TipoMoneda
        search_text = "TipoMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TipoMoneda').text = value
                print(f'TipoMoneda : {value}')
                break
    
        # Write TipoCambio
        search_text = "TipoCambio" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TipoCambio').text = value
                print(f'TipoCambio : {value}')
                break
    
        # Write MontoGravadoTotalOtraMoneda
        search_text = "MontoGravadoTotalOtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'MontoGravadoTotalOtraMoneda').text = value
                print(f'MontoGravadoTotalOtraMoneda : {value}')
                break
    
        # Write MontoGravado3OtraMoneda
        search_text = "MontoGravado3OtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'MontoGravado3OtraMoneda').text = value
                print(f'MontoGravado3OtraMoneda : {value}')
                break
    
        # Write TotalITBISOtraMoneda
        search_text = "TotalITBISOtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TotalITBISOtraMoneda').text = value
                print(f'TotalITBISOtraMoneda : {value}')
                break

        # Write TotalITBIS3OtraMoneda
        search_text = "TotalITBIS3OtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TotalITBIS3OtraMoneda').text = value
                print(f'TotalITBIS3OtraMoneda : {value}')
                break

        # Write MontoTotalOtraMoneda
        search_text = "MontoTotalOtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'MontoTotalOtraMoneda').text = value
                print(f'MontoTotalOtraMoneda : {value}')
                break

        # """
        DetallesItems = ET.SubElement(root, 'DetallesItems')

        item_count = 62
        search_text = "NumeroLinea[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                col_index = col
                while True:
                    item_count -= 1
                    if item_count < 0 : 
                        break
                    Item = ET.SubElement(DetallesItems, 'Item')

                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'NumeroLinea').text = value
                    print(f'NumeroLinea : {value}')
                    col_index += 1
                    

                    # TablaCodigosItem
                    TablaCodigosItem = ET.SubElement(Item, 'TablaCodigosItem')
                    CodigosItem_count = 5;
                    while True : 
                        CodigosItem_count -= 1
                        if CodigosItem_count < 0 : 
                            break

                        CodigosItem = ET.SubElement(TablaCodigosItem, 'CodigosItem')
                        
                        # TipoCodigo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(CodigosItem, 'TipoCodigo').text = value
                        print(f'TipoCodigo : {value}')
                        col_index += 1

                        # CodigoItem
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(CodigosItem, 'CodigoItem').text = value
                        print(f'CodigoItem : {value}')
                        col_index += 1

                    # IndicadorFacturacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'IndicadorFacturacion').text = value
                    print(f'IndicadorFacturacion : {value}')
                    col_index += 1

                    Retencion = ET.SubElement(Item, 'Retencion')
                    # IndicadorAgenteRetencionoPercepcion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Retencion, 'IndicadorAgenteRetencionoPercepcion').text = value
                    print(f'IndicadorAgenteRetencionoPercepcion : {value}')
                    col_index += 1
                    
                    # MontoITBISRetenido
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Retencion, 'MontoITBISRetenido').text = value
                    print(f'MontoITBISRetenido : {value}')
                    col_index += 1

                    # MontoISRRetenido
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Retencion, 'MontoISRRetenido').text = value
                    print(f'MontoISRRetenido : {value}')
                    col_index += 1

                    # NombreItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'NombreItem').text = value
                    print(f'NombreItem : {value}')
                    col_index += 1

                    # IndicadorBienoServicio
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'IndicadorBienoServicio').text = value
                    print(f'IndicadorBienoServicio : {value}')
                    col_index += 1

                    # DescripcionItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'DescripcionItem').text = value
                    print(f'DescripcionItem : {value}')
                    col_index += 1

                    # CantidadItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'CantidadItem').text = value
                    print(f'CantidadItem : {value}')
                    col_index += 1

                    # UnidadMedida
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'UnidadMedida').text = value
                    print(f'UnidadMedida : {value}')
                    col_index += 1

                    # CantidadReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value != "#e" :
                        ET.SubElement(Item, 'CantidadReferencia').text = value
                        print(f'CantidadReferencia : {value}')
                    col_index += 1

                    # UnidadReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value != "#e" :
                        ET.SubElement(Item, 'UnidadReferencia').text = value
                        print(f'UnidadReferencia : {value}')
                    col_index += 1

                    # TablaSubcantidad = ET.SubElement(Item, 'TablaSubcantidad')

                    SubcantidadItem_count = 5
                    while True : 
                        SubcantidadItem_count -= 1
                        if SubcantidadItem_count < 0 :
                            break

                        # SubcantidadItem = ET.SubElement(TablaSubcantidad, 'SubcantidadItem')
                        
                        # Subcantidad
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value != "#e" :
                            # ET.SubElement(SubcantidadItem, 'Subcantidad').text = value
                            print(f'Subcantidad : {value}')
                        col_index += 1
                        
                        # CodigoSubcantidad
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value != "#e" :
                            # ET.SubElement(SubcantidadItem, 'CodigoSubcantidad').text = value
                            print(f'CodigoSubcantidad : {value}')
                        col_index += 1

                    # GradosAlcohol
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value != "#e" :
                        ET.SubElement(Item, 'GradosAlcohol').text = value
                        print(f'GradosAlcohol : {value}')
                    col_index += 1

                    # PrecioUnitarioReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value != "#e" :
                        ET.SubElement(Item, 'PrecioUnitarioReferencia').text = value
                        print(f'PrecioUnitarioReferencia : {value}')
                    col_index += 1

                    # FechaElaboracion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value != "#e" :
                        ET.SubElement(Item, 'FechaElaboracion').text = value
                        print(f'FechaElaboracion : {value}')
                    col_index += 1

                    # FechaVencimientoItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'FechaVencimientoItem').text = value
                    print(f'FechaVencimientoItem : {value}')
                    col_index += 1

                    
                    Mineria = ET.SubElement(Item, 'Mineria')
                    
                    # PesoNetoKilogramo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Mineria, 'PesoNetoKilogramo').text = value
                    print(f'PesoNetoKilogramo : {value}')
                    col_index += 1
                    
                    # PesoNetoMineria
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Mineria, 'PesoNetoMineria').text = value
                    print(f'PesoNetoMineria : {value}')
                    col_index += 1

                    # TipoAfiliacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Mineria, 'TipoAfiliacion').text = value
                    print(f'TipoAfiliacion : {value}')
                    col_index += 1

                    # Liquidacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Mineria, 'Liquidacion').text = value
                    print(f'Liquidacion : {value}')
                    col_index += 1

                    
                    # PrecioUnitarioItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(Item, 'PrecioUnitarioItem').text = value
                    print(f'PrecioUnitarioItem : {value}')
                    col_index += 1

                    # DescuentoMonto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'DescuentoMonto').text = value
                    print(f'DescuentoMonto : {value}')
                    col_index += 1

                    TablaSubDescuento = ET.SubElement(Item, 'TablaSubDescuento')

                    SubDescuento_count = 5

                    while True : 
                        SubDescuento_count -= 1
                        if SubDescuento_count < 0 : 
                            break

                        #SubDescuento
                        SubDescuento = ET.SubElement(TablaSubDescuento, 'SubDescuento')

                        # TipoSubDescuento
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubDescuento, 'TipoSubDescuento').text = value
                        print(f'TipoSubDescuento : {value}')
                        col_index += 1

                        # SubDescuentoPorcentaje
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubDescuento, 'SubDescuentoPorcentaje').text = value
                        print(f'SubDescuentoPorcentaje : {value}')
                        col_index += 1

                        # MontoSubDescuento
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubDescuento, 'MontoSubDescuento').text = value
                        print(f'MontoSubDescuento : {value}')
                        col_index += 1

                    # RecargoMonto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'RecargoMonto').text = value
                    print(f'RecargoMonto : {value}')
                    col_index += 1

                    TablaSubRecargo = ET.SubElement(Item, 'TablaSubRecargo')

                    SubRecargo_count = 5

                    while True :
                        SubRecargo_count -= 1
                        if SubRecargo_count < 0 :
                            break    
                        SubRecargo = ET.SubElement(TablaSubRecargo, 'SubRecargo')

                        # TipoSubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubRecargo, 'TipoSubRecargo').text = value
                        print(f'TipoSubRecargo : {value}')
                        col_index += 1

                        # SubRecargoPorcentaje
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubRecargo, 'SubRecargoPorcentaje').text = value
                        print(f'SubRecargoPorcentaje : {value}')
                        col_index += 1

                        
                        # MontosubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(SubRecargo, 'MontosubRecargo').text = value
                        print(f'MontosubRecargo : {value}')
                        col_index += 1


                    # TablaImpuestoAdicional = ET.SubElement(Item, 'TablaImpuestoAdicional')

                    ImpuestoAdicional_count = 2
                    while True :
                        ImpuestoAdicional_count -= 1
                        if ImpuestoAdicional_count < 0 : 
                            break
                        # ImpuestoAdicional =  ET.SubElement(TablaImpuestoAdicional, 'ImpuestoAdicional')                                   
                        # MontosubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value != "#e" :
                            ET.SubElement(ImpuestoAdicional, 'TipoImpuesto').text = value
                            print(f'TipoImpuesto : {value}')
                        col_index += 1               
    
                    OtraMonedaDetalle = ET.SubElement(Item, 'OtraMonedaDetalle')

                    # PrecioOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'PrecioOtraMoneda').text = value
                    print(f'PrecioOtraMoneda : {value}')
                    col_index += 1

                    # DescuentoOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'DescuentoOtraMoneda').text = value
                    print(f'DescuentoOtraMoneda : {value}')
                    col_index += 1

                    # RecargoOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'RecargoOtraMoneda').text = value
                    print(f'RecargoOtraMoneda : {value}')

                    col_index += 1
                    # MontoItemOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'MontoItemOtraMoneda').text = value
                    print(f'MontoItemOtraMoneda : {value}')
                    col_index += 1
                    
                    # MontoItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'MontoItem').text = value
                    print(f'MontoItem : {value}')
                    col_index += 1
        
        Subtotales = ET.SubElement(root, 'Subtotales')
        Subtotal = ET.SubElement(Subtotales, 'Subtotal')

        # Write NumeroSubTotal
        search_text = "NumeroSubTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'NumeroSubTotal').text = value
                print(f'Version : {cell_value}')
                break
        
        # Write DescripcionSubtotal
        search_text = "DescripcionSubtotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'DescripcionSubtotal').text = value
                print(f'DescripcionSubtotal : {cell_value}')
                break
        
        # Write Orden
        search_text = "Orden" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'Orden').text = value
                print(f'Orden : {cell_value}')
                break
        
        # Write SubTotalMontoGravadoTotal
        search_text = "SubTotalMontoGravadoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'SubTotalMontoGravadoTotal').text = value
                print(f'SubTotalMontoGravadoTotal : {cell_value}')
                break
    
        # Write SubTotalMontoGravadoI3
        search_text = "SubTotalMontoGravadoI3" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'SubTotalMontoGravadoI3').text = value
                print(f'SubTotalMontoGravadoI3 : {cell_value}')
                break
    
        # Write SubTotaITBIS
        search_text = "SubTotaITBIS" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'SubTotaITBIS').text = value
                print(f'SubTotaITBIS : {cell_value}')
                break

        # Write SubTotaITBIS3
        search_text = "SubTotaITBIS3" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'SubTotaITBIS3').text = value
                print(f'SubTotaITBIS3 : {cell_value}')
                break

        # Write MontoSubTotal
        search_text = "MontoSubTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'MontoSubTotal').text = value
                print(f'MontoSubTotal : {cell_value}')
                break

        # Write Lineas
        search_text = "Lineas" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'Lineas').text = value
                print(f'Lineas : {cell_value}')
                break

        DescuentosORecargos = ET.SubElement(root, 'DescuentosORecargos')
        DescuentoORecargo_count = 2

        search_text = "NumeroLineaDoR[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            
            if cell_value == search_text:
                col_index = col
                while True : 
                    DescuentoORecargo_count -= 1
                    if DescuentoORecargo_count < 0:
                        break

                    DescuentoORecargo = ET.SubElement(DescuentosORecargos, 'DescuentoORecargo')

                    # NumeroLinea
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'NumeroLinea').text = value
                    print(f'NumeroLinea : {cell_value}')
                    col_index +=1

                    # TipoAjuste
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'TipoAjuste').text = value
                    print(f'TipoAjuste : {cell_value}')
                    col_index +=1
                    
                    # DescripcionDescuentooRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'DescripcionDescuentooRecargo').text = value
                    print(f'DescripcionDescuentooRecargo : {cell_value}')
                    col_index +=1
                                            
                    # TipoValor
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'TipoValor').text = value
                    print(f'TipoValor : {cell_value}')
                    col_index +=1
                                                        
                    # ValorDescuentooRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'ValorDescuentooRecargo').text = value
                    print(f'ValorDescuentooRecargo : {cell_value}')
                    col_index +=1
                                                                
                    # MontoDescuentooRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'MontoDescuentooRecargo').text = value
                    print(f'MontoDescuentooRecargo : {cell_value}')
                    col_index +=1
                                                                            
                    # MontoDescuentooRecargoOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'MontoDescuentooRecargoOtraMoneda').text = value
                    print(f'MontoDescuentooRecargoOtraMoneda : {cell_value}')
                    col_index +=1        

                    # IndicadorFacturacionDescuentooRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(DescuentoORecargo, 'IndicadorFacturacionDescuentooRecargo').text = value
                    print(f'IndicadorFacturacionDescuentooRecargo : {cell_value}')         
        

        Paginacion = ET.SubElement(root, 'Paginacion')
        Pagina_count = 2

        search_text = "PaginaNo[1]" 
        col_index : int
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:

                col_index = col
                    
                while True:
                    Pagina_count -= 1
                    if Pagina_count < 0:
                        break

                    Pagina = ET.SubElement(Paginacion, 'Pagina')

                    # PaginNo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'PaginaNo').text = value
                    print(f'PaginaNo : {cell_value}')
                    print(f'PaginaNo col_index : {col_index}')
                    col_index += 1

                    # NoLineaDesde
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'NoLineaDesde').text = value
                    print(f'NoLineaDesde : {cell_value}')
                    col_index += 1

                    # NoLineaHasta
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'NoLineaHasta').text = value
                    print(f'NoLineaHasta : {value}')
                    col_index += 1
                    
                    # SubtotalMontoGravadoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'SubtotalMontoGravadoPagina').text = value
                    print(f'SubtotalMontoGravadoPagina : {value}')
                    col_index += 1
                                    
                    # SubtotalMontoGravado1Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravado1Pagina').text = value
                    print(f'SubtotalMontoGravado1Pagina : {value}')
                    col_index += 1
                                                    
                    # SubtotalMontoGravado2Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravado2Pagina').text = value
                    print(f'SubtotalMontoGravado2Pagina : {value}')
                    col_index += 1
                                                    
                    # SubtotalMontoGravado3Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'SubtotalMontoGravado3Pagina').text = value
                    print(f'SubtotalMontoGravado3Pagina : {value}')
                    col_index += 1
                                                                    
                    # SubtotalExentoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalExentoPagina').text = value
                    print(f'SubtotalExentoPagina : {value}')
                    col_index += 1
                                                                
                    # SubtotalItbisPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'SubtotalItbisPagina').text = value
                    print(f'SubtotalItbisPagina : {value}')
                    col_index += 1
                                                                        
                    # SubtotalItbis1Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbis1Pagina').text = value
                    print(f'SubtotalItbis1Pagina : {value}')
                    col_index += 1
                                                                                        
                    # SubtotalItbis2Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbis2Pagina').text = value
                    print(f'SubtotalItbis2Pagina : {value}')
                    col_index += 1
                                                                                            
                    # SubtotalItbis3Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'SubtotalItbis3Pagina').text = value
                    print(f'SubtotalItbis3Pagina : {value}')
                    col_index += 1
                                                                                                            
                    # SubtotalImpuestoAdicionalPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalImpuestoAdicionalPagina').text = value
                    print(f'SubtotalImpuestoAdicionalPagina : {value}')
                    col_index += 1

                    SubtotalImpuestoAdicional = ET.SubElement(Pagina, 'SubtotalImpuestoAdicional')
                    
                    # SubtotalImpuestoSelectivoConsumoEspecificoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalImpuestoSelectivoConsumoEspecificoPagina').text =value
                    print(f'SubtotalImpuestoSelectivoConsumoEspecificoPagina : {value}')
                    col_index += 1

                    # SubtotalOtrosImpuesto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalOtrosImpuesto').text =value
                    print(f'SubtotalOtrosImpuesto : {value}')
                    col_index += 1

                    # MontoSubtotalPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    ET.SubElement(Pagina, 'MontoSubtotalPagina').text =value
                    print(f'MontoSubtotalPagina : {value}')
                    col_index += 1

                    # SubtotalMontoNoFacturablePagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    ET.SubElement(Pagina, 'SubtotalMontoNoFacturablePagina').text =value
                    print(f'SubtotalMontoNoFacturablePagina : {value}')
                    print(f'SubtotalMontoNoFacturablePagina _ index : {col_index}')
                    col_index += 1

        InformacionReferencia = ET.SubElement(root, 'InformacionReferencia')
        
        # NCFModificado
        col_index += 1
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'NCFModificado').text =value
        print(f'NCFModificado : {value}')
        col_index += 1
            
        # RNCOtroContribuyente
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'RNCOtroContribuyente').text =value
        print(f'RNCOtroContribuyente : {value}')
        col_index += 1
            
        # FechaNCFModificado
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'FechaNCFModificado').text =value
        print(f'FechaNCFModificado : {value}')
        col_index += 1
            
        # CodigoModificacion
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'CodigoModificacion').text =value
        print(f'CodigoModificacion : {value}')
        col_index += 1

        ET.SubElement(root, 'FechaHoraFirma').text ="08-09-2025 09:47:51"
       
        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')
        pretty_xml_str = pretty_xml_as_str.decode('utf-8') if isinstance(pretty_xml_as_str, bytes) else pretty_xml_as_str

        cleaned_xml = clean_xml_safe(pretty_xml_as_str)
        print(cleaned_xml)

        self.invoice_name = f'{self.RNCEmisor}{self.ENCF}'
        path = os.path.join(os.path.dirname(__file__), f'data/{self.invoice_name}.xml')
        
        with open(path, 'wb') as f:
            f.write(cleaned_xml.encode('utf-8'))
        
        xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
        return cleaned_xml