
import openpyxl
import xml.etree.ElementTree as ET
import logging
import os
from xml.dom import minidom

workbook = openpyxl.load_workbook("test_data.xlsx")
sheet = workbook["RFCE"]

class RFCEClient : 
    def __init__(self) :
       self.invoice_name = ""
       self.RNCEmisor = ""
       self.ENCF = ""
       self._logger =  logging.getLogger(__name__)

    def create_rfce_xml(self, in_row : int) :

        """Generate DGII-compliant XML from an Odoo invoice."""
        # Create root element with namespaces
        root = ET.Element('RFCE', {
                'xmlns:xs': 'http://www.w3.org/2001/XMLSchema'
            })
        
        encabezado = ET.SubElement(root, 'Encabezado')

        # Read Version
        search_text = "Version" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(encabezado, 'Version').text= str(sheet.cell(in_row, column=col).value)
                
        id_doc = ET.SubElement(encabezado, 'IdDoc')

        # Read TipoeCF
        search_text = "TipoeCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(id_doc, 'TipoeCF').text = str(sheet.cell(in_row, column=col).value)

        # Read eNCF
        search_text = "ENCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                ET.SubElement(id_doc, 'eNCF').text = value
                self.ENCF = value
                self._logger.info(f'ENCF : {value}')

        # Read TipoIngresos
        search_text = "TipoIngresos" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(id_doc, 'TipoIngresos').text= str(sheet.cell(in_row, column=col).value)

        # Read TipoPago
        search_text = "TipoPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(id_doc, 'TipoPago').text = str(sheet.cell(in_row, column=col).value)

        # Read FormaDePago
        # TablaFormasPago = ET.SubElement(id_doc, 'TablaFormasPago')

        # search_text = "FormaPago[1]"
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         FormaDePago_count = 7
        #         col_index = col
        #         while True :
        #             FormaDePago_count -= 1
        #             if FormaDePago_count < 0 :
        #                 break            
        #             FormaDePago = ET.SubElement(TablaFormasPago, 'FormaDePago')

        #             value = str(sheet.cell(in_row, column= col_index ).value)
        #             if value == "#e" :
        #                 value = ""
        #             ET.SubElement(FormaDePago, 'FormaPago').text = value
        #             col_index += 1

        #             value = str(sheet.cell(in_row, column= col_index).value)
        #             if value == "#e" :
        #                 value = ""
        #             ET.SubElement(FormaDePago, 'MontoPago').text = value
        #             col_index += 1
        
        Emisor = ET.SubElement(encabezado, 'Emisor')
        
        # RNCEmisor
        search_text = "RNCEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                self.RNCEmisor = value
                ET.SubElement(Emisor, 'RNCEmisor').text = value
                self._logger.info(f'RNCEmisor : {value}')
                break

        # Write RazonSocialEmisor
        search_text = "RazonSocialEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RazonSocialEmisor').text = value
                self._logger.info(f'RazonSocialEmisor : {value}')
                break
    
        # Write FechaEmision
        search_text = "FechaEmision" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'FechaEmision').text = value
                self._logger.info(f'FechaEmision : {value}')
                break

        Comprador = ET.SubElement(encabezado, 'Comprador')

        # Write RNCComprador
        search_text = "RNCComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'RNCComprador').text = value
                self._logger.info(f'RNCComprador : {value}')
                break

    # Write IdentificadorExtranjero
        # search_text = "IdentificadorExtranjero" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Comprador, 'IdentificadorExtranjero').text = value
        #         __logger.info(f'IdentificadorExtranjero : {value}')
        #         break

        # Write RazonSocialComprador
        search_text = "RazonSocialComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Comprador, 'RazonSocialComprador').text = value
                self._logger.info(f'RazonSocialComprador : {value}')
                break
        
        Totales = ET.SubElement(encabezado, 'Totales')
        
        # Write MontoGravadoTotal
        search_text = "MontoGravadoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoGravadoTotal').text = value
                self._logger.info(f'MontoGravadoTotal : {value}')
                break
    
        # Write MontoGravadoI1
        search_text = "MontoGravadoI1" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoGravadoI1').text = value
                self._logger.info(f'MontoGravadoI1 : {value}')
                break
    
        # Write MontoGravadoI2
        # search_text = "MontoGravadoI2" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoGravadoI2').text = value
        #         __logger.info(f'MontoGravadoI2 : {value}')
        #         break
    
        # Write MontoGravadoI3
        # search_text = "MontoGravadoI3" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoGravadoI3').text = value
        #         __logger.info(f'MontoGravadoI3 : {value}')
        #         break
    
        # Write MontoExento
        # search_text = "MontoExento" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoExento').text = value
        #         __logger.info(f'MontoExento : {value}')
        #         break
        
        # Write TotalITBIS
        search_text = "TotalITBIS" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'TotalITBIS').text = value
                self._logger.info(f'TotalITBIS : {value}')
                break
    
        # Write TotalITBIS1
        search_text = "TotalITBIS1" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'TotalITBIS1').text = value
                self._logger.info(f'TotalITBIS1 : {value}')
                break
        """
        # Write TotalITBIS2
        # search_text = "TotalITBIS2" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'TotalITBIS2').text = value
        #         __logger.info(f'TotalITBIS2 : {value}')
        #         break

        # Write TotalITBIS3
        # search_text = "TotalITBIS3" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'TotalITBIS3').text = value
        #         __logger.info(f'TotalITBIS3 : {value}')
        #         break

        # Write MontoImpuestoAdicional
        # search_text = "MontoImpuestoAdicional" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoImpuestoAdicional').text = value
        #         __logger.info(f'MontoImpuestoAdicional : {value}')
        #         break

        ImpuestosAdicionales = ET.SubElement(Totales, 'ImpuestosAdicionales')
        search_text = "TipoImpuesto[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:

                col_index = col
                ImpuestoAdicional_count = 4
                while True:
                    ImpuestoAdicional_count -= 1
                    if ImpuestoAdicional_count < 0:
                        break

                    ImpuestoAdicional = ET.SubElement(ImpuestosAdicionales, 'ImpuestoAdicional')

                    # TipoImpuesto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(ImpuestoAdicional, 'TipoImpuesto').text = value
                    __logger.info(f'TipoImpuesto : {value}')
                    col_index +=1

                    # # TasaImpuestoAdicional
                    # value = str(sheet.cell(in_row, column=col_index).value)
                    # if value == "#e" :
                    #     value = "" 
                    # ET.SubElement(ImpuestoAdicional, 'TasaImpuestoAdicional').text = value
                    # __logger.info(f'TasaImpuestoAdicional : {value}')
                    # col_index +=1

                    # MontoImpuestoSelectivoConsumoEspecifico
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(ImpuestoAdicional, 'MontoImpuestoSelectivoConsumoEspecifico').text = value
                    __logger.info(f'MontoImpuestoSelectivoConsumoEspecifico : {value}')
                    col_index +=1

                    # MontoImpuestoSelectivoConsumoAdvalorem
                    value = str(sheet.cell(in_row, column=col_index + 3).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(ImpuestoAdicional, 'MontoImpuestoSelectivoConsumoAdvalorem').text = value
                    __logger.info(f'MontoImpuestoSelectivoConsumoAdvalorem : {value}')
                    col_index +=1


                    # OtrosImpuestosAdicionales
                    value = str(sheet.cell(in_row, column=col_index + 4).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(ImpuestoAdicional, 'OtrosImpuestosAdicionales').text = value
                    __logger.info(f'OtrosImpuestosAdicionales : {value}')
                    col_index +=1
        """

        # Write MontoTotal
        search_text = "MontoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoTotal').text = value
                self._logger.info(f'MontoTotal : {value}')
                break

        # Write MontoNoFacturable
        # search_text = "MontoNoFacturable" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoNoFacturable').text = value
        #         __logger.info(f'MontoNoFacturable : {value}')
        #         break

        # # Write MontoPeriodo
        # search_text = "MontoPeriodo" 
        # for col in range(1, sheet.max_column + 1):
        #     cell_value = sheet.cell(1, column=col).value
        #     if cell_value == search_text:
        #         value = str(sheet.cell(in_row, column=col).value)
        #         if value == "#e" :
        #             value = "" 

        #         ET.SubElement(Totales, 'MontoPeriodo').text = value
        #         __logger.info(f'MontoPeriodo : {value}')
        #         break

        # Read CodigoSeguridadeCF
        search_text = "CodigoSeguridadeCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "000000"
                else :
                    value = value[:6].ljust(6, "0")

                ET.SubElement(encabezado, 'CodigoSeguridadeCF').text = value
                self._logger.info(f'CodigoSeguridadeCF : {value}')

                # Convert to XML string
        xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
        
        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')

        self.invoice_name = f"{self.RNCEmisor}{self.ENCF}"

        path = os.path.join(os.path.dirname(__file__), f'data/{self.invoice_name}.xml')
        
        with open(path, 'wb') as f:
            f.write(pretty_xml_as_str)

        self._logger.info("Finished the creating RFCE xml")

        # Convert to lxml for signing
        return xml_str