import json
import os
from typing import List
from xml.dom import minidom
import openpyxl
import xml.etree.ElementTree as ET
from dataclasses import dataclass
import requests

from logger import __logger
from dgii_client import DGIICFService
from rfce import RFCEClient
from ecf31 import ECF31
from ecf32 import ECF32
from ecf33 import ECF33
from ecf34 import ECF34
from ecf41 import ECF41
from ecf43 import ECF43
from ecf44 import ECF44
from ecf45 import ECF45
from ecf46 import ECF46

workbook = openpyxl.load_workbook("test_data.xlsx")
sheet = workbook["ECF"]

@dataclass
class Param:
    RNCEmisor: str = ""
    ENCF: str = ""
    CodigoSeguridadeCF: str = ""

param = Param()
invoice_name = ""

def generate_dummy_dgii_xml():
    """Generate fully compliant DGII dummy XML for immediate client demo"""
    root = ET.Element('ECF', {
        'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        'xmlns:ds': 'http://www.w3.org/2000/09/xmldsig#',
        'xsi:schemaLocation': 'https://ecf.dgii.gov.do/esquemas/ecf/1.1'
    })

    # 1. Encabezado (Header) - Correct structure
    encabezado = ET.SubElement(root, 'Encabezado')
    
    # Required elements
    ET.SubElement(encabezado, 'Version').text = '1.1'
    
    id_doc = ET.SubElement(encabezado, 'IdDoc')
    ET.SubElement(id_doc, 'TipoeCF').text = "32"
    ET.SubElement(id_doc, 'eNCF').text = "E320000000011"
    ET.SubElement(id_doc, 'IndicadorMontoGravado').text = '1'  # 1=Yes, 0=No
    ET.SubElement(id_doc, 'FechaLimitePago').text = '2025-08-30'  # Optional
    # ET.SubElement(id_doc, 'TipoPago').text = ''  # "Contado" or "Crédito"
    ET.SubElement(id_doc, 'TerminoPago').text = ''  # "Contado" or "Crédito"
    TablaFormasPago = ET.SubElement(id_doc, 'TablaFormasPago')  # "Contado" or "Crédito"
    FormaDePago = ET.SubElement(TablaFormasPago, 'FormaDePago')
    ET.SubElement(FormaDePago, 'FormaPago').text = '01'  # Cash
    ET.SubElement(FormaDePago, 'MontoPago').text = '1180.00'  # Total amount paid    

    FormaDePago = ET.SubElement(TablaFormasPago, 'FormaDePago')
    ET.SubElement(FormaDePago, 'FormaPago').text = ''  # Cash
    ET.SubElement(FormaDePago, 'MontoPago').text = ''  # Total amount paid    

    
    emisor = ET.SubElement(encabezado, 'Emisor')
    ET.SubElement(emisor, 'RNCEmisor').text = '132641566'
    ET.SubElement(emisor, 'RazonSocialEmisor').text = 'EMPRESA DEMO SRL'
    ET.SubElement(emisor, 'NombreComercial').text = 'DEMO COMERCIAL'
    
    # Required address elements
    ET.SubElement(emisor, 'DireccionEmisor').text = 'Calle Principal #123'
    ET.SubElement(emisor, 'Municipio').text = 'Distrito Nacional'
    ET.SubElement(emisor, 'Provincia').text = 'Santo Domingo'
    
    # Contact information
    telefono = ET.SubElement(emisor, 'TablaTelefonoEmisor')
    
    # First phone number (landline)
    # telefono1 = ET.SubElement(telefono, 'TelefonoEmisor')
    # ET.SubElement(telefono1, 'NumeroTelefono').text = '8095551234'  # No hyphens
    # ET.SubElement(telefono1, 'TipoTelefono').text = '1'  # 1 for landline

    ET.SubElement(telefono, 'TelefonoEmisor').text = '8495555678'  # No hyphens

    ET.SubElement(emisor, 'CorreoEmisor').text = 'info@empresademo.com'
    ET.SubElement(emisor, 'WebSite').text = 'www.empresademo.com'
    
    # Economic information
    ET.SubElement(emisor, 'ActividadEconomica').text = 'VENTA AL POR MENOR'
    ET.SubElement(emisor, 'CodigoVendedor').text = 'VEN001'
    
    # Optional but recommended
    ET.SubElement(emisor, 'NumeroFacturaInterna').text = 'FAC-1001'
    ET.SubElement(emisor, 'NumeroPedidoInterno').text = 'PED-5001'
    ET.SubElement(emisor, 'ZonaVenta').text = 'ZONA 1'
    ET.SubElement(emisor, 'RutaVenta').text = 'RUTA A'
    
    # Additional info
    info_adicional = ET.SubElement(emisor, 'InformacionAdicionalEmisor')
    ET.SubElement(info_adicional, 'InformacionAdicional', {
        'nombre': 'Sucursal',
        'texto': 'Principal'
    })

    ET.SubElement(emisor, 'FechaEmision').text = "10-10-2020"
    
    comprador = ET.SubElement(encabezado, 'Comprador')
    ET.SubElement(comprador, 'RNCComprador').text = '987654321'
    ET.SubElement(comprador, 'RazonSocialComprador').text = 'CLIENTE DEMOSTRACION'

    # Critical fix: Correct order of elements
    # info_adicional = ET.SubElement(encabezado, 'InformacionesAdicionales')
    # ET.SubElement(info_adicional, 'InformacionAdicional', {
    #     'nombre': 'Observaciones',
    #     'texto': 'DEMO PARA CLIENTE - NO ES FACTURA REAL'
    # })
    
    ET.SubElement(encabezado, 'Transporte')  # Required empty element

    if not encabezado.find('Totales'):  # Only add if not present
        totales = ET.SubElement(encabezado, 'Totales')
        ET.SubElement(totales, 'MontoTotal').text = '1180.00'  # Required
        ET.SubElement(totales, 'ValorPagar').text = '1180.00'  # Required
        ET.SubElement(totales, 'TotalITBISRetenido').text = '0.00'  # Required for B2B
        ET.SubElement(totales, 'TotalISRRetencion').text = '0.00'  # Required for services
        ET.SubElement(totales, 'TotalITBISPercepcion').text = '0.00'  # Optional

    # 2. Detalles (Line Items)
    detalles = ET.SubElement(root, 'Detalles')
    
    # Sample product 1
    item1 = ET.SubElement(detalles, 'Item')
    ET.SubElement(item1, 'Descripcion').text = 'SERVICIO DE DEMOSTRACION'
    ET.SubElement(item1, 'Cantidad').text = '1'
    ET.SubElement(item1, 'PrecioUnitario').text = '1000.00'
    ET.SubElement(item1, 'ITBIS').text = '180.00'  # 18%
    ET.SubElement(item1, 'MontoItem').text = '1180.00'

    # 3. Totales (Must be direct child of ECF)
    totales = ET.SubElement(root, 'Totales')
    ET.SubElement(totales, 'MontoTotal').text = '1180.00'
    ET.SubElement(totales, 'ITBISTotal').text = '180.00'

    rough_string = ET.tostring(root, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')

    path = os.path.join(os.path.dirname(__file__), 'data/row_invoice.xml')
    
    with open(path, 'wb') as f:
        f.write(pretty_xml_as_str)

    # Return formatted XML
    xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
    return xml_str

__trackId : str
__token : str

def create_e_cf_47(in_row : int) :

    """Generate DGII-compliant XML from an Odoo invoice."""
    # Create root element with namespaces
    root = ET.Element('ECF', {
         'xmlns:xs' : "http://www.w3.org/2001/XMLSchema",
    })

    # 1. Encabezado
    encabezado = ET.SubElement(root, 'Encabezado')

    # Write Version
    search_text = "Version" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value != "#e" :    
                ET.SubElement(encabezado, 'Version').text = '1.1'
                __logger.info(f'Version : {value}')
                __logger.info(f'Version Col Index : {col}')
            break

    # IdDoc
    id_doc = ET.SubElement(encabezado, 'IdDoc')

    # Write TipoeCF
    search_text = "TipoeCF" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 
            ET.SubElement(id_doc, 'TipoeCF').text = value
            __logger.info(f'TipoeCF : {cell_value}')
            break

    # Write eNCF
    search_text = "ENCF" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 
            ET.SubElement(id_doc, 'eNCF').text = value
            __logger.info(f'eNCF : {cell_value}')
            break

    # Write FechaVencimientoSecuencia
    search_text = "FechaVencimientoSecuencia" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'FechaVencimientoSecuencia').text = value
            __logger.info(f'FechaVencimientoSecuencia : {value}')
            break

    # Write TipoPago
    search_text = "TipoPago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'TipoPago').text = value
            __logger.info(f'TipoPago : {value}')
            break

    # Write FechaLimitePago
    search_text = "FechaLimitePago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'FechaLimitePago').text = value
            __logger.info(f'FechaLimitePago : {value}')
            break


    # Write TerminoPago
    search_text = "TerminoPago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'TerminoPago').text = value
            __logger.info(f'TerminoPago : {value}')
            break
    
    
    TablaFormasPago = ET.SubElement(id_doc, 'TablaFormasPago')

    search_text = "FormaPago[1]"
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            FormaDePago_count = 7
            col_index = col
            while True :
                FormaDePago_count -= 1
                if FormaDePago_count < 0 :
                    break            
                FormaDePago = ET.SubElement(TablaFormasPago, 'FormaDePago')

                value = str(sheet.cell(in_row, column= col_index ).value)
                if value == "#e" :
                    value = ""
                ET.SubElement(FormaDePago, 'FormaPago').text = value
                col_index += 1

                value = str(sheet.cell(in_row, column= col_index).value)
                if value == "#e" :
                    value = ""
                ET.SubElement(FormaDePago, 'MontoPago').text = value
                col_index += 1
    
    # Write TipoCuentaPago
    search_text = "TipoCuentaPago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'TipoCuentaPago').text = value
            __logger.info(f'TipoCuentaPago : {value}')
            break
        
    # Write NumeroCuentaPago
    search_text = "NumeroCuentaPago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'NumeroCuentaPago').text = value
            __logger.info(f'NumeroCuentaPago : {value}')
            break
        
    # Write BancoPago
    search_text = "BancoPago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'BancoPago').text = value
            __logger.info(f'BancoPago : {value}')
            break

    # Write FechaDesde
    search_text = "FechaDesde" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'FechaDesde').text = value
            __logger.info(f'FechaDesde : {value}')
            break    

    # Write FechaHasta
    search_text = "FechaHasta" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'FechaHasta').text = value
            __logger.info(f'FechaHasta : {value}')
            break    

    # Write TotalPaginas
    search_text = "TotalPaginas" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(id_doc, 'TotalPaginas').text = value
            __logger.info(f'TotalPaginas : {value}')
            break

    Emisor = ET.SubElement(encabezado, 'Emisor')

    # Write RNCEmisor
    search_text = "RNCEmisor" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'RNCEmisor').text = value
            __logger.info(f'RNCEmisor : {value}')
            break

    # Write RazonSocialEmisor
    search_text = "RazonSocialEmisor" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'RazonSocialEmisor').text = value
            __logger.info(f'RazonSocialEmisor : {value}')
            break

    # Write NombreComercial
    search_text = "NombreComercial" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'NombreComercial').text = value
            __logger.info(f'NombreComercial : {value}')
            break

    # Write Sucursal
    search_text = "Sucursal" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'Sucursal').text = value
            __logger.info(f'Sucursal : {value}')
            break

    # Write DireccionEmisor
    search_text = "DireccionEmisor" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'DireccionEmisor').text = value
            __logger.info(f'DireccionEmisor : {value}')
            break

    # Write Municipio
    search_text = "Municipio" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'Municipio').text = value
            __logger.info(f'Municipio : {value}')
            break

    # Write Provincia
    search_text = "Provincia" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'Provincia').text = value
            __logger.info(f'Provincia : {value}')
            break

    TablaTelefonoEmisor = ET.SubElement(Emisor, 'TablaTelefonoEmisor')

    # Write TelefonoEmisor
    search_text = "TelefonoEmisor[1]" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            TelefonoEmisor_count = 3
            col_index = col
            while True :
                TelefonoEmisor_count -= 1
                if TelefonoEmisor_count < 0 : 
                    break

                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""
                ET.SubElement(TablaTelefonoEmisor, 'TelefonoEmisor').text = value           
                __logger.info(f'TelefonoEmisor : {value}')

                col_index +=1
             
    # Write CorreoEmisor
    search_text = "CorreoEmisor" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'CorreoEmisor').text = value
            __logger.info(f'CorreoEmisor : {value}')
            break

    # Write WebSite
    search_text = "WebSite" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'WebSite').text = value
            __logger.info(f'WebSite : {value}')
            break

    # Write ActividadEconomica
    search_text = "ActividadEconomica" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'ActividadEconomica').text = value
            __logger.info(f'ActividadEconomica : {value}')
            break

    # Write NumeroFacturaInterna
    search_text = "NumeroFacturaInterna" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'NumeroFacturaInterna').text = value
            __logger.info(f'NumeroFacturaInterna : {value}')
            break
 
    # Write NumeroPedidoInterno
    search_text = "NumeroPedidoInterno" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'NumeroPedidoInterno').text = value
            __logger.info(f'NumeroPedidoInterno : {value}')
            break

    # Write InformacionAdicionalEmisor
    search_text = "InformacionAdicionalEmisor" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'InformacionAdicionalEmisor').text = value
            __logger.info(f'InformacionAdicionalEmisor : {value}')
            break

    # Write FechaEmision
    search_text = "FechaEmision" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Emisor, 'FechaEmision').text = value
            __logger.info(f'FechaEmision : {value}')
            break

    Comprador = ET.SubElement(encabezado, 'Comprador')

    # Write IdentificadorExtranjero
    search_text = "IdentificadorExtranjero" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Comprador, 'IdentificadorExtranjero').text = value
            __logger.info(f'IdentificadorExtranjero : {value}')
            break

    # Write RazonSocialComprador
    search_text = "RazonSocialComprador" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Comprador, 'RazonSocialComprador').text = value
            __logger.info(f'RazonSocialComprador : {value}')
            break
    
    # Transporte
    Transporte = ET.SubElement(encabezado, 'Transporte')

    # Write PaisDestino
    search_text = "PaisDestino" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Transporte, 'PaisDestino').text = value
            __logger.info(f'PaisDestino : {value}')
            break

    Totales = ET.SubElement(encabezado, 'Totales')
   
    # Write MontoExento
    search_text = "MontoExento" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'MontoExento').text = value
            __logger.info(f'MontoExento : {value}')
            break

    # Write MontoTotal
    search_text = "MontoTotal" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'MontoTotal').text = value
            __logger.info(f'MontoTotal : {value}')
            break

    # Write MontoPeriodo
    search_text = "MontoPeriodo" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'MontoPeriodo').text = value
            __logger.info(f'MontoPeriodo : {value}')
            break

    # Write SaldoAnterior
    search_text = "SaldoAnterior" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'SaldoAnterior').text = value
            __logger.info(f'SaldoAnterior : {value}')
            break

    # Write MontoAvancePago
    search_text = "MontoAvancePago" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'MontoAvancePago').text = value
            __logger.info(f'MontoAvancePago : {value}')
            break

    # Write ValorPagar
    search_text = "ValorPagar" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'ValorPagar').text = value
            __logger.info(f'ValorPagar : {value}')
            break

    # Write TotalISRRetencion
    search_text = "TotalISRRetencion" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(Totales, 'TotalISRRetencion').text = value
            __logger.info(f'TotalISRRetencion : {value}')
            break

    OtraMoneda = ET.SubElement(encabezado, 'OtraMoneda')
  
    # Write TipoMoneda
    search_text = "TipoMoneda" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(OtraMoneda, 'TipoMoneda').text = value
            __logger.info(f'TipoMoneda : {value}')
            break
  
    # Write TipoCambio
    search_text = "TipoCambio" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(OtraMoneda, 'TipoCambio').text = value
            __logger.info(f'TipoCambio : {value}')
            break
  
    # Write MontoExentoOtraMoneda
    search_text = "MontoExentoOtraMoneda" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(OtraMoneda, 'MontoExentoOtraMoneda').text = value
            __logger.info(f'MontoExentoOtraMoneda : {value}')
            break
  
    # Write MontoTotalOtraMoneda
    search_text = "MontoTotalOtraMoneda" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = "" 

            ET.SubElement(OtraMoneda, 'MontoTotalOtraMoneda').text = value
            __logger.info(f'MontoTotalOtraMoneda : {value}')
            break

    # """
    DetallesItems = ET.SubElement(root, 'DetallesItems')

    item_count = 62
    search_text = "NumeroLinea[1]" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            col_index = col
            while True:
                item_count -= 1
                if item_count < 0 : 
                    break
                Item = ET.SubElement(DetallesItems, 'Item')

                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'NumeroLinea').text = value
                __logger.info(f'NumeroLinea : {value}')
                col_index += 1
                

                # TablaCodigosItem
                TablaCodigosItem = ET.SubElement(Item, 'TablaCodigosItem')
                CodigosItem_count = 5;
                while True : 
                    CodigosItem_count -= 1
                    if CodigosItem_count < 0 : 
                        break

                    CodigosItem = ET.SubElement(TablaCodigosItem, 'CodigosItem')
                    
                    # TipoCodigo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(CodigosItem, 'TipoCodigo').text = value
                    __logger.info(f'TipoCodigo : {value}')
                    col_index += 1

                    # CodigoItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(CodigosItem, 'CodigoItem').text = value
                    __logger.info(f'CodigoItem : {value}')
                    col_index += 1

                # IndicadorFacturacion
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'IndicadorFacturacion').text = value
                __logger.info(f'IndicadorFacturacion : {value}')
                col_index += 1

                Retencion = ET.SubElement(Item, 'Retencion')
                # IndicadorAgenteRetencionoPercepcion
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Retencion, 'IndicadorAgenteRetencionoPercepcion').text = value
                __logger.info(f'IndicadorAgenteRetencionoPercepcion : {value}')
                col_index += 1
                
                # MontoITBISRetenido
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Retencion, 'MontoITBISRetenido').text = value
                __logger.info(f'MontoITBISRetenido : {value}')
                col_index += 1

                # MontoISRRetenido
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Retencion, 'MontoISRRetenido').text = value
                __logger.info(f'MontoISRRetenido : {value}')
                col_index += 1

                # NombreItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'NombreItem').text = value
                __logger.info(f'NombreItem : {value}')
                col_index += 1

                # IndicadorBienoServicio
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'IndicadorBienoServicio').text = value
                __logger.info(f'IndicadorBienoServicio : {value}')
                col_index += 1

                # DescripcionItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'DescripcionItem').text = value
                __logger.info(f'DescripcionItem : {value}')
                col_index += 1

                # CantidadItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'CantidadItem').text = value
                __logger.info(f'CantidadItem : {value}')
                col_index += 1

                # UnidadMedida
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'UnidadMedida').text = value
                __logger.info(f'UnidadMedida : {value}')
                col_index += 1

                # CantidadReferencia
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'CantidadReferencia').text = value
                __logger.info(f'CantidadReferencia : {value}')
                col_index += 1

                # UnidadReferencia
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'UnidadReferencia').text = value
                __logger.info(f'UnidadReferencia : {value}')
                col_index += 1

                # TablaSubcantidad = ET.SubElement(Item, 'TablaSubcantidad')

                SubcantidadItem_count = 5
                while True : 
                    SubcantidadItem_count -= 1
                    if SubcantidadItem_count < 0 :
                        break

                    # SubcantidadItem = ET.SubElement(TablaSubcantidad, 'SubcantidadItem')
                    
                    # Subcantidad
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubcantidadItem, 'Subcantidad').text = value
                    __logger.info(f'Subcantidad : {value}')
                    col_index += 1
                    
                    # CodigoSubcantidad
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubcantidadItem, 'CodigoSubcantidad').text = value
                    __logger.info(f'CodigoSubcantidad : {value}')
                    col_index += 1

                # GradosAlcohol
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'GradosAlcohol').text = value
                __logger.info(f'GradosAlcohol : {value}')
                col_index += 1

                # PrecioUnitarioReferencia
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'PrecioUnitarioReferencia').text = value
                __logger.info(f'PrecioUnitarioReferencia : {value}')
                col_index += 1

                # FechaElaboracion
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'FechaElaboracion').text = value
                __logger.info(f'FechaElaboracion : {value}')
                col_index += 1

                # FechaVencimientoItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'FechaVencimientoItem').text = value
                __logger.info(f'FechaVencimientoItem : {value}')
                col_index += 1

                
                # Mineria = ET.SubElement(Item, 'Mineria')
                
                # PesoNetoKilogramo
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Mineria, 'PesoNetoKilogramo').text = value
                __logger.info(f'PesoNetoKilogramo : {value}')
                col_index += 1
                
                # PesoNetoMineria
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Mineria, 'PesoNetoMineria').text = value
                __logger.info(f'PesoNetoMineria : {value}')
                col_index += 1

                # TipoAfiliacion
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Mineria, 'TipoAfiliacion').text = value
                __logger.info(f'TipoAfiliacion : {value}')
                col_index += 1

                # Liquidacion
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Mineria, 'Liquidacion').text = value
                __logger.info(f'Liquidacion : {value}')
                col_index += 1

                
                # PrecioUnitarioItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 
                ET.SubElement(Item, 'PrecioUnitarioItem').text = value
                __logger.info(f'PrecioUnitarioItem : {value}')
                col_index += 1

                # DescuentoMonto
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'DescuentoMonto').text = value
                __logger.info(f'DescuentoMonto : {value}')
                col_index += 1

                # TablaSubDescuento = ET.SubElement(Item, 'TablaSubDescuento')

                SubDescuento_count = 5

                while True : 
                    SubDescuento_count -= 1
                    if SubDescuento_count < 0 : 
                        break

                    #SubDescuento
                    # SubDescuento = ET.SubElement(TablaSubDescuento, 'SubDescuento')

                    # TipoSubDescuento
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubDescuento, 'TipoSubDescuento').text = value
                    __logger.info(f'TipoSubDescuento : {value}')
                    col_index += 1

                    # SubDescuentoPorcentaje
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubDescuento, 'SubDescuentoPorcentaje').text = value
                    __logger.info(f'SubDescuentoPorcentaje : {value}')
                    col_index += 1

                    # MontoSubDescuento
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubDescuento, 'MontoSubDescuento').text = value
                    __logger.info(f'MontoSubDescuento : {value}')
                    col_index += 1


                # RecargoMonto
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                # ET.SubElement(Item, 'RecargoMonto').text = value
                __logger.info(f'RecargoMonto : {value}')
                col_index += 1

                # TablaSubRecargo = ET.SubElement(Item, 'TablaSubRecargo')

                SubRecargo_count = 5

                while True :
                    SubRecargo_count -= 1
                    if SubRecargo_count < 0 :
                        break    
                    # SubRecargo = ET.SubElement(TablaSubRecargo, 'SubRecargo')

                    # TipoSubRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubRecargo, 'TipoSubRecargo').text = value
                    __logger.info(f'TipoSubRecargo : {value}')
                    col_index += 1

                    # SubRecargoPorcentaje
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubRecargo, 'SubRecargoPorcentaje').text = value
                    __logger.info(f'SubRecargoPorcentaje : {value}')
                    col_index += 1

                    
                    # MontosubRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(SubRecargo, 'MontosubRecargo').text = value
                    __logger.info(f'MontosubRecargo : {value}')
                    col_index += 1

                # TablaImpuestoAdicional = ET.SubElement(Item, 'TablaImpuestoAdicional')

                ImpuestoAdicional_count = 2
                while True :
                    ImpuestoAdicional_count -= 1
                    if ImpuestoAdicional_count < 0 : 
                        break
                    # ImpuestoAdicional =  ET.SubElement(TablaImpuestoAdicional, 'ImpuestoAdicional')                                   
                    # MontosubRecargo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(ImpuestoAdicional, 'TipoImpuesto').text = value
                    __logger.info(f'TipoImpuesto : {value}')
                    col_index += 1
                
                OtraMonedaDetalle = ET.SubElement(Item, 'OtraMonedaDetalle')

                # PrecioOtraMoneda
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMonedaDetalle, 'PrecioOtraMoneda').text = value
                __logger.info(f'PrecioOtraMoneda : {value}')
                col_index += 1

                # DescuentoOtraMoneda
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMonedaDetalle, 'DescuentoOtraMoneda').text = value
                __logger.info(f'DescuentoOtraMoneda : {value}')
                col_index += 1

                # RecargoOtraMoneda
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMonedaDetalle, 'RecargoOtraMoneda').text = value
                __logger.info(f'RecargoOtraMoneda : {value}')

                col_index += 1
                # MontoItemOtraMoneda
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMonedaDetalle, 'MontoItemOtraMoneda').text = value
                __logger.info(f'MontoItemOtraMoneda : {value}')
                col_index += 1
                
                # MontoItem
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Item, 'MontoItem').text = value
                __logger.info(f'MontoItem : {value}')
                col_index += 1
    
    Subtotales = ET.SubElement(root, 'Subtotales')
    Subtotal = ET.SubElement(Subtotales, 'Subtotal')

    # Write NumeroSubTotal
    search_text = "NumeroSubTotal" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'NumeroSubTotal').text = value
            __logger.info(f'Version : {cell_value}')
            break
    
    # Write DescripcionSubtotal
    search_text = "DescripcionSubtotal" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'DescripcionSubtotal').text = value
            __logger.info(f'DescripcionSubtotal : {cell_value}')
            break
    
    # Write Orden
    search_text = "Orden" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'Orden').text = value
            __logger.info(f'Orden : {cell_value}')
            break
    
    # Write SubTotalExento
    search_text = "SubTotalExento" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'SubTotalExento').text = value
            __logger.info(f'SubTotalExento : {cell_value}')
            break

    # Write MontoSubTotal
    search_text = "MontoSubTotal" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'MontoSubTotal').text = value
            __logger.info(f'MontoSubTotal : {cell_value}')
            break

    # Write Lineas
    search_text = "Lineas" 
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:
            value = str(sheet.cell(in_row, column=col).value)
            if value == "#e" :
                value = ""             
            ET.SubElement(Subtotal, 'Lineas').text = value
            __logger.info(f'Lineas : {cell_value}')
            break

    Paginacion = ET.SubElement(root, 'Paginacion')
    Pagina_count = 2

    search_text = "PaginaNo[1]" 
    col_index : int
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(1, column=col).value
        if cell_value == search_text:

            col_index = col
                
            while True:
                Pagina_count -= 1
                if Pagina_count < 0:
                    break

                Pagina = ET.SubElement(Paginacion, 'Pagina')

                # PaginNo
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Pagina, 'PaginaNo').text = value
                __logger.info(f'PaginaNo : {cell_value}')
                __logger.info(f'PaginaNo col_index : {col_index}')
                col_index += 1

                # NoLineaDesde
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Pagina, 'NoLineaDesde').text = value
                __logger.info(f'NoLineaDesde : {cell_value}')
                col_index += 1

                # NoLineaHasta
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Pagina, 'NoLineaHasta').text = value
                __logger.info(f'NoLineaHasta : {value}')
                col_index += 1
                
                # SubtotalMontoGravadoPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalMontoGravadoPagina').text = value
                __logger.info(f'SubtotalMontoGravadoPagina : {value}')
                col_index += 1
                                
                # SubtotalMontoGravado1Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalMontoGravado1Pagina').text = value
                __logger.info(f'SubtotalMontoGravado1Pagina : {value}')
                col_index += 1
                                                
                # SubtotalMontoGravado2Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalMontoGravado2Pagina').text = value
                __logger.info(f'SubtotalMontoGravado2Pagina : {value}')
                col_index += 1
                                                   
                # SubtotalMontoGravado3Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalMontoGravado3Pagina').text = value
                __logger.info(f'SubtotalMontoGravado3Pagina : {value}')
                col_index += 1
                                                                   
                # SubtotalExentoPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Pagina, 'SubtotalExentoPagina').text = value
                __logger.info(f'SubtotalExentoPagina : {value}')
                col_index += 1
                                                               
                # SubtotalItbisPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalItbisPagina').text = value
                __logger.info(f'SubtotalItbisPagina : {value}')
                col_index += 1
                                                                      
                # SubtotalItbis1Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalItbis1Pagina').text = value
                __logger.info(f'SubtotalItbis1Pagina : {value}')
                col_index += 1
                                                                                      
                # SubtotalItbis2Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalItbis2Pagina').text = value
                __logger.info(f'SubtotalItbis2Pagina : {value}')
                col_index += 1
                                                                                         
                # SubtotalItbis3Pagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalItbis3Pagina').text = value
                __logger.info(f'SubtotalItbis3Pagina : {value}')
                col_index += 1
                                                                                                         
                # SubtotalImpuestoAdicionalPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""             
                # ET.SubElement(Pagina, 'SubtotalImpuestoAdicionalPagina').text = value
                __logger.info(f'SubtotalImpuestoAdicionalPagina : {value}')
                col_index += 1

                # SubtotalImpuestoAdicional = ET.SubElement(Pagina, 'SubtotalImpuestoAdicional')
                
                # SubtotalImpuestoSelectivoConsumoEspecificoPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""     
                # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalImpuestoSelectivoConsumoEspecificoPagina').text =value
                __logger.info(f'SubtotalImpuestoSelectivoConsumoEspecificoPagina : {value}')
                col_index += 1

                # SubtotalOtrosImpuesto
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""     
                # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalOtrosImpuesto').text =value
                __logger.info(f'SubtotalOtrosImpuesto : {value}')
                col_index += 1

                # MontoSubtotalPagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""     
                ET.SubElement(Pagina, 'MontoSubtotalPagina').text =value
                __logger.info(f'MontoSubtotalPagina : {value}')
                col_index += 1

                # SubtotalMontoNoFacturablePagina
                value = str(sheet.cell(in_row, column=col_index).value)
                if value == "#e" :
                    value = ""     
                # ET.SubElement(Pagina, 'SubtotalMontoNoFacturablePagina').text =value
                __logger.info(f'SubtotalMontoNoFacturablePagina : {value}')
                __logger.info(f'SubtotalMontoNoFacturablePagina _ index : {col_index}')
                col_index += 1

    InformacionReferencia = ET.SubElement(root, 'InformacionReferencia')
    
    # NCFModificado
    col_index += 1
    value = str(sheet.cell(in_row, column=col_index).value)
    if value == "#e" :
        value = ""     
    ET.SubElement(InformacionReferencia, 'NCFModificado').text =value
    __logger.info(f'NCFModificado : {value}')
    col_index += 1
        
    # RNCOtroContribuyente
    value = str(sheet.cell(in_row, column=col_index).value)
    if value == "#e" :
        value = ""     
    ET.SubElement(InformacionReferencia, 'RNCOtroContribuyente').text =value
    __logger.info(f'RNCOtroContribuyente : {value}')
    col_index += 1
        
    # FechaNCFModificado
    value = str(sheet.cell(in_row, column=col_index).value)
    if value == "#e" :
        value = ""     
    ET.SubElement(InformacionReferencia, 'FechaNCFModificado').text =value
    __logger.info(f'FechaNCFModificado : {value}')
    col_index += 1
        
    # CodigoModificacion
    value = str(sheet.cell(in_row, column=col_index).value)
    if value == "#e" :
        value = ""     
    ET.SubElement(InformacionReferencia, 'CodigoModificacion').text =value
    __logger.info(f'CodigoModificacion : {value}')
    col_index += 1

    rough_string = ET.tostring(root, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')

    path = os.path.join(os.path.dirname(__file__), 'data/row_invoice.xml')
    
    with open(path, 'wb') as f:
        f.write(pretty_xml_as_str)
    
    xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
    return xml_str

def main():

    try:
        dgii_service = DGIICFService(dgii_env='prod', company_id=1)

        # Step 1: Get the seed
        semilla = dgii_service.get_semilla()
        if not semilla:
            raise  __logger.error("Failed to retrieve semilla from DGII.")

        __logger.info(f"Received semilla: {semilla}")

        # Step 2: Sign the semilla
        signed_semilla = dgii_service.sign_semilla(semilla)

        # __logger.info(f"Signed semilla: {signed_semilla}")

        # Step 3: Validate the signed semilla and get the token
        token = dgii_service.validate_semilla()
        if not token:
            raise __logger.error("Failed to validate semilla with DGII and obtain token.")

        __logger.info(f"Received token: {token}")

        """ filetest for debuging"""
        # xml_file_path = os.path.join(os.path.dirname(__file__), 'data/row_invoice.xml')

        # with open(xml_file_path, 'r', encoding='utf-8') as file:
        #     xml_str = file.read()
        # Print or use the XML string
        # print(xml_str)

        # Step 4: Generate and sign e-CF XML     
        # xml_str = generate_dummy_dgii_xml()  
        # xml_str = create_e_cf_47(26) # 25, 26
        # __logger.info(f"invoice xml: {xml_str}")


        #################################
        ###########   RFCE   ############ 
        #################################

        # rfce_client = RFCEClient()
        # xml_str = rfce_client.create_rfce_xml(5)

        # invoice_name = rfce_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")

        
        ##############################################
        ###########        E-CF-31        ############
        ###########   row : 7, 8, 9, 10   ############
        ##############################################
        # e_cf_31_client = ECF31()
        # xml_str = e_cf_31_client.create_e_cf_31(7)

        # invoice_name = e_cf_31_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")

        
        ##############################################
        ######        E-CF-32                  #######
        ######   row : 11, 12, 13, 14, 15, 3   #######
        ##############################################
        # e_cf_32_client = ECF32()
        # xml_str = e_cf_32_client.create_e_cf_32(3)

        # invoice_name = e_cf_32_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")

        
        ##############################################
        ######           E-CF-33               #######
        ######           row : 2               #######
        ##############################################
        # e_cf_33_client = ECF33()
        # xml_str = e_cf_33_client.create_e_cf_33(2)

        # invoice_name = e_cf_33_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")

                
        ##############################################
        ######           E-CF-34               #######
        ######           row : 4, 5            #######
        ##############################################
        # e_cf_34_client = ECF34()
        # xml_str = e_cf_34_client.create_e_cf_34(4)

        # invoice_name = e_cf_34_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")
        
                
        ##############################################
        ######           E-CF-41               #######
        ######           row : 6, 16           #######
        ##############################################
        # e_cf_41_client = ECF41()
        # xml_str = e_cf_41_client.create_e_cf_41(6)

        # invoice_name = e_cf_41_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")

        ##############################################
        ######           E-CF-43               #######
        ######           row : 17, 18          #######
        ##############################################
        # e_cf_43_client = ECF43()
        # xml_str = e_cf_43_client.create_e_cf_43(17)

        # invoice_name = e_cf_43_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")


        ##############################################
        ######           E-CF-44               #######
        ######           row : 19, 20          #######
        ##############################################
        # e_cf_44_client = ECF44()
        # xml_str = e_cf_44_client.create_e_cf_44(19)

        # invoice_name = e_cf_44_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")


        ##############################################
        ######           E-CF-45               #######
        ######           row : 21, 22          #######
        ##############################################
        # e_cf_45_client = ECF45()
        # xml_str = e_cf_45_client.create_e_cf_45(19)

        # invoice_name = e_cf_45_client.invoice_name
        # __logger.info(f"Invoice Name: {invoice_name}")


        ##############################################
        ######           E-CF-46               #######
        ######           row : 23, 24          #######
        ##############################################
        e_cf_46_client = ECF46()
        xml_str = e_cf_46_client.create_e_cf_46(23)

        invoice_name = e_cf_46_client.invoice_name
        __logger.info(f"Invoice Name: {invoice_name}")

        
        signed_xml = dgii_service.sign_xml(xml_str, invoice_name)
        # __logger.info(f"signed invoice xml: {signed_xml}")

        # Step 5: Submit the signed e-CF XML to DGII using the token
        # response = dgii_service.submit_rfce(token)
        response = dgii_service.submit_ecf(token)
        trackId = ""
        response_data = json.loads(response.content.decode('utf-8'))

        __logger.info(f"Submitting Response:  {response_data}")

        # Log the response and update the invoice status
        if response.status_code == 200:
            response_data = json.loads(response.content.decode('utf-8'))
            __logger.info(f"Submitted to DGII, response: trackId:  {response_data['trackId']}")
            trackId = response_data['trackId']
        else:
            __logger.error(f"Error submitting invoice to DGII: {response.text}")
            return

        # Step 6: Track the status of e-CF to DGII using the token
        response = dgii_service.track_ecf(trackId, token)
        # response = dgii_service.track_rfce(token, param)
        __logger.info(f"track result: {response.content}")

        response_data = json.loads(response.content.decode('utf-8'))
        __logger.info(f"dgii_track_id {trackId}, dgii_token : {token}")

    except Exception as e:
        __logger.error(f"Error processing invoice: {str(e)}")

if __name__ == "__main__":
    main()