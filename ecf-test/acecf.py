import openpyxl
import xml.etree.ElementTree as ET
import logging
import os
from xml.dom import minidom

workbook = openpyxl.load_workbook("acecf.xlsx")
sheet = workbook["ACEECF_Generadas"]

class ACECF : 
    def __init__(self) :
       self.invoice_name = ""
       self.RNCComprador = ""
       self.ENCF = ""

    def create_acecf_xml(self, in_row : int) :

        """Generate DGII-compliant XML from an Odoo invoice."""
        # Create root element with namespaces
        root = ET.Element('ACECF', {
                'xmlns:xs': 'http://www.w3.org/2001/XMLSchema'
            })
        
        DetalleAprobacionComercial = ET.SubElement(root, 'DetalleAprobacionComercial')

        # Read Version
        search_text = "Version" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'Version').text= str(sheet.cell(in_row, column=col).value)
                break
        
        # Read RNCEmisor
        search_text = "RNCEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'RNCEmisor').text= str(sheet.cell(in_row, column=col).value)
                break

        # Read eNCF
        search_text = "eNCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                val = str(sheet.cell(in_row, column=col).value)
                ET.SubElement(DetalleAprobacionComercial, 'eNCF').text= val
                self.ENCF = val
                break
   
        # Read FechaEmision
        search_text = "FechaEmision" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'FechaEmision').text= str(sheet.cell(in_row, column=col).value)
                break
   
        # Read MontoTotal
        search_text = "MontoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'MontoTotal').text= str(sheet.cell(in_row, column=col).value)
                break
       
        # Read RNCComprador
        search_text = "RNCComprador" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                val = str(sheet.cell(in_row, column=col).value)
                ET.SubElement(DetalleAprobacionComercial, 'RNCComprador').text= val
                self.RNCComprador = val
                break  
               
        # Read Estado
        search_text = "Estado" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'Estado').text= str(sheet.cell(in_row, column=col).value)
                break
                           
        # Read FechaHoraAprobacionComercial
        search_text = "FechaHoraAprobacionComercial" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                ET.SubElement(DetalleAprobacionComercial, 'FechaHoraAprobacionComercial').text= str(sheet.cell(in_row, column=col).value)
                break

        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')

        self.invoice_name = f'{self.RNCComprador}{self.ENCF}'
        path = os.path.join(os.path.dirname(__file__), f'data/{self.invoice_name}.xml')
        
        with open(path, 'wb') as f:
            f.write(pretty_xml_as_str)
        
        xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
        return xml_str