import openpyxl
import xml.etree.ElementTree as ET
import logging
import os
from xml.dom import minidom

workbook = openpyxl.load_workbook("test_data.xlsx")
sheet = workbook["ECF"]

class ECF43:
    """Class to handle E-CF-34 XML generation."""
    def __init__(self):
        self.RNCEmisor = ""
        self.ENCF = ""
        self.RNCComprador = ""
        self.invoice_name = ""
    
    def create_e_cf_43(self, in_row : int) :

        """Generate DGII-compliant XML from an Odoo invoice."""
        # Create root element with namespaces
        root = ET.Element('ECF', {
            'xmlns:xs' : "http://www.w3.org/2001/XMLSchema",
        })

        # 1. Encabezado
        encabezado = ET.SubElement(root, 'Encabezado')

        # Write Version
        search_text = "Version" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value != "#e" :    
                    ET.SubElement(encabezado, 'Version').text = '1.1'
                    print(f'Version : {value}')
                    print(f'Version Col Index : {col}')
                break

        # IdDoc
        id_doc = ET.SubElement(encabezado, 'IdDoc')

        # Write TipoeCF
        search_text = "TipoeCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 
                ET.SubElement(id_doc, 'TipoeCF').text = value
                print(f'TipoeCF : {cell_value}')
                break

        # Write eNCF
        search_text = "ENCF" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 
                ET.SubElement(id_doc, 'eNCF').text = value
                self.ENCF = value
                print(f'eNCF : {cell_value}')
                break

        # Write FechaVencimientoSecuencia
        search_text = "FechaVencimientoSecuencia" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'FechaVencimientoSecuencia').text = value
                print(f'FechaVencimientoSecuencia : {value}')
                break

        # Write TipoPago
        search_text = "TipoPago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TipoPago').text = value
                print(f'TipoPago : {value}')
                break

        # Write TotalPaginas
        search_text = "TotalPaginas" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(id_doc, 'TotalPaginas').text = value
                print(f'TotalPaginas : {value}')
                break

        Emisor = ET.SubElement(encabezado, 'Emisor')

        # Write RNCEmisor
        search_text = "RNCEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RNCEmisor').text = value
                self.RNCEmisor = value
                print(f'RNCEmisor : {value}')
                break

        # Write RazonSocialEmisor
        search_text = "RazonSocialEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'RazonSocialEmisor').text = value
                print(f'RazonSocialEmisor : {value}')
                break

        # Write NombreComercial
        search_text = "NombreComercial" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NombreComercial').text = value
                print(f'NombreComercial : {value}')
                break

        # Write Sucursal
        search_text = "Sucursal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Sucursal').text = value
                print(f'Sucursal : {value}')
                break

        # Write DireccionEmisor
        search_text = "DireccionEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'DireccionEmisor').text = value
                print(f'DireccionEmisor : {value}')
                break

        # Write Municipio
        search_text = "Municipio" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Municipio').text = value
                print(f'Municipio : {value}')
                break

        # Write Provincia
        search_text = "Provincia" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'Provincia').text = value
                print(f'Provincia : {value}')
                break

        TablaTelefonoEmisor = ET.SubElement(Emisor, 'TablaTelefonoEmisor')

        # Write TelefonoEmisor
        search_text = "TelefonoEmisor[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                TelefonoEmisor_count = 3
                col_index = col
                while True :
                    TelefonoEmisor_count -= 1
                    if TelefonoEmisor_count < 0 : 
                        break

                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""
                    ET.SubElement(TablaTelefonoEmisor, 'TelefonoEmisor').text = value           
                    print(f'TelefonoEmisor : {value}')

                    col_index +=1
                
        # Write CorreoEmisor
        search_text = "CorreoEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'CorreoEmisor').text = value
                print(f'CorreoEmisor : {value}')
                break

        # Write WebSite
        search_text = "WebSite" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'WebSite').text = value
                print(f'WebSite : {value}')
                break

        # Write ActividadEconomica
        search_text = "ActividadEconomica" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'ActividadEconomica').text = value
                print(f'ActividadEconomica : {value}')
                break

        # Write NumeroFacturaInterna
        search_text = "NumeroFacturaInterna" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NumeroFacturaInterna').text = value
                print(f'NumeroFacturaInterna : {value}')
                break
    
        # Write NumeroPedidoInterno
        search_text = "NumeroPedidoInterno" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'NumeroPedidoInterno').text = value
                print(f'NumeroPedidoInterno : {value}')
                break
    
        # Write InformacionAdicionalEmisor
        search_text = "InformacionAdicionalEmisor" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'InformacionAdicionalEmisor').text = value
                print(f'InformacionAdicionalEmisor : {value}')
                break

        # Write FechaEmision
        search_text = "FechaEmision" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Emisor, 'FechaEmision').text = value
                print(f'FechaEmision : {value}')
                break

        Totales = ET.SubElement(encabezado, 'Totales')
    
        # Write MontoExento
        search_text = "MontoExento" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoExento').text = value
                print(f'MontoExento : {value}')
                break

        # Write MontoTotal
        search_text = "MontoTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoTotal').text = value
                print(f'MontoTotal : {value}')
                break

        # Write MontoPeriodo
        search_text = "MontoPeriodo" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoPeriodo').text = value
                print(f'MontoPeriodo : {value}')
                break

        # Write SaldoAnterior
        search_text = "SaldoAnterior" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'SaldoAnterior').text = value
                print(f'SaldoAnterior : {value}')
                break

        # Write MontoAvancePago
        search_text = "MontoAvancePago" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'MontoAvancePago').text = value
                print(f'MontoAvancePago : {value}')
                break

        # Write ValorPagar
        search_text = "ValorPagar" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(Totales, 'ValorPagar').text = value
                print(f'ValorPagar : {value}')
                break

        OtraMoneda = ET.SubElement(encabezado, 'OtraMoneda')
    
        # Write TipoMoneda
        search_text = "TipoMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TipoMoneda').text = value
                print(f'TipoMoneda : {value}')
                break
    
        # Write TipoCambio
        search_text = "TipoCambio" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'TipoCambio').text = value
                print(f'TipoCambio : {value}')
                break
    
        # Write MontoExentoOtraMoneda
        search_text = "MontoExentoOtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'MontoExentoOtraMoneda').text = value
                print(f'MontoExentoOtraMoneda : {value}')
                break
    
        # Write MontoTotalOtraMoneda
        search_text = "MontoTotalOtraMoneda" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = "" 

                ET.SubElement(OtraMoneda, 'MontoTotalOtraMoneda').text = value
                print(f'MontoTotalOtraMoneda : {value}')
                break

        # """
        DetallesItems = ET.SubElement(root, 'DetallesItems')

        item_count = 62
        search_text = "NumeroLinea[1]" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                col_index = col
                while True:
                    item_count -= 1
                    if item_count < 0 : 
                        break
                    Item = ET.SubElement(DetallesItems, 'Item')

                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'NumeroLinea').text = value
                    print(f'NumeroLinea : {value}')
                    col_index += 1
                    

                    # TablaCodigosItem
                    TablaCodigosItem = ET.SubElement(Item, 'TablaCodigosItem')
                    CodigosItem_count = 5;
                    while True : 
                        CodigosItem_count -= 1
                        if CodigosItem_count < 0 : 
                            break

                        CodigosItem = ET.SubElement(TablaCodigosItem, 'CodigosItem')
                        
                        # TipoCodigo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(CodigosItem, 'TipoCodigo').text = value
                        print(f'TipoCodigo : {value}')
                        col_index += 1

                        # CodigoItem
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        ET.SubElement(CodigosItem, 'CodigoItem').text = value
                        print(f'CodigoItem : {value}')
                        col_index += 1

                    # IndicadorFacturacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'IndicadorFacturacion').text = value
                    print(f'IndicadorFacturacion : {value}')
                    col_index += 1

                    # Retencion = ET.SubElement(Item, 'Retencion')
                    # IndicadorAgenteRetencionoPercepcion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Retencion, 'IndicadorAgenteRetencionoPercepcion').text = value
                    print(f'IndicadorAgenteRetencionoPercepcion : {value}')
                    col_index += 1
                    
                    # MontoITBISRetenido
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Retencion, 'MontoITBISRetenido').text = value
                    print(f'MontoITBISRetenido : {value}')
                    col_index += 1

                    # MontoISRRetenido
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Retencion, 'MontoISRRetenido').text = value
                    print(f'MontoISRRetenido : {value}')
                    col_index += 1

                    # NombreItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'NombreItem').text = value
                    print(f'NombreItem : {value}')
                    col_index += 1

                    # IndicadorBienoServicio
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'IndicadorBienoServicio').text = value
                    print(f'IndicadorBienoServicio : {value}')
                    col_index += 1

                    # DescripcionItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'DescripcionItem').text = value
                    print(f'DescripcionItem : {value}')
                    col_index += 1

                    # CantidadItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'CantidadItem').text = value
                    print(f'CantidadItem : {value}')
                    col_index += 1

                    # UnidadMedida
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'UnidadMedida').text = value
                    print(f'UnidadMedida : {value}')
                    col_index += 1

                    # CantidadReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'CantidadReferencia').text = value
                    print(f'CantidadReferencia : {value}')
                    col_index += 1

                    # UnidadReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'UnidadReferencia').text = value
                    print(f'UnidadReferencia : {value}')
                    col_index += 1

                    # TablaSubcantidad = ET.SubElement(Item, 'TablaSubcantidad')

                    SubcantidadItem_count = 5
                    while True : 
                        SubcantidadItem_count -= 1
                        if SubcantidadItem_count < 0 :
                            break

                        # SubcantidadItem = ET.SubElement(TablaSubcantidad, 'SubcantidadItem')
                        
                        # Subcantidad
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubcantidadItem, 'Subcantidad').text = value
                        print(f'Subcantidad : {value}')
                        col_index += 1
                        
                        # CodigoSubcantidad
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubcantidadItem, 'CodigoSubcantidad').text = value
                        print(f'CodigoSubcantidad : {value}')
                        col_index += 1

                    # GradosAlcohol
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'GradosAlcohol').text = value
                    print(f'GradosAlcohol : {value}')
                    col_index += 1

                    # PrecioUnitarioReferencia
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'PrecioUnitarioReferencia').text = value
                    print(f'PrecioUnitarioReferencia : {value}')
                    col_index += 1

                    # FechaElaboracion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'FechaElaboracion').text = value
                    print(f'FechaElaboracion : {value}')
                    col_index += 1

                    # FechaVencimientoItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'FechaVencimientoItem').text = value
                    print(f'FechaVencimientoItem : {value}')
                    col_index += 1

                    
                    # Mineria = ET.SubElement(Item, 'Mineria')
                    
                    # PesoNetoKilogramo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Mineria, 'PesoNetoKilogramo').text = value
                    print(f'PesoNetoKilogramo : {value}')
                    col_index += 1
                    
                    # PesoNetoMineria
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Mineria, 'PesoNetoMineria').text = value
                    print(f'PesoNetoMineria : {value}')
                    col_index += 1

                    # TipoAfiliacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Mineria, 'TipoAfiliacion').text = value
                    print(f'TipoAfiliacion : {value}')
                    col_index += 1

                    # Liquidacion
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Mineria, 'Liquidacion').text = value
                    print(f'Liquidacion : {value}')
                    col_index += 1

                    
                    # PrecioUnitarioItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 
                    ET.SubElement(Item, 'PrecioUnitarioItem').text = value
                    print(f'PrecioUnitarioItem : {value}')
                    col_index += 1

                    # DescuentoMonto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'DescuentoMonto').text = value
                    print(f'DescuentoMonto : {value}')
                    col_index += 1

                    # TablaSubDescuento = ET.SubElement(Item, 'TablaSubDescuento')

                    SubDescuento_count = 5

                    while True : 
                        SubDescuento_count -= 1
                        if SubDescuento_count < 0 : 
                            break

                        #SubDescuento
                        # SubDescuento = ET.SubElement(TablaSubDescuento, 'SubDescuento')

                        # TipoSubDescuento
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubDescuento, 'TipoSubDescuento').text = value
                        print(f'TipoSubDescuento : {value}')
                        col_index += 1

                        # SubDescuentoPorcentaje
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubDescuento, 'SubDescuentoPorcentaje').text = value
                        print(f'SubDescuentoPorcentaje : {value}')
                        col_index += 1

                        # MontoSubDescuento
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubDescuento, 'MontoSubDescuento').text = value
                        print(f'MontoSubDescuento : {value}')
                        col_index += 1


                    # RecargoMonto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    # ET.SubElement(Item, 'RecargoMonto').text = value
                    print(f'RecargoMonto : {value}')
                    col_index += 1

                    # TablaSubRecargo = ET.SubElement(Item, 'TablaSubRecargo')

                    SubRecargo_count = 5

                    while True :
                        SubRecargo_count -= 1
                        if SubRecargo_count < 0 :
                            break    
                        # SubRecargo = ET.SubElement(TablaSubRecargo, 'SubRecargo')

                        # TipoSubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubRecargo, 'TipoSubRecargo').text = value
                        print(f'TipoSubRecargo : {value}')
                        col_index += 1

                        # SubRecargoPorcentaje
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubRecargo, 'SubRecargoPorcentaje').text = value
                        print(f'SubRecargoPorcentaje : {value}')
                        col_index += 1

                        
                        # MontosubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(SubRecargo, 'MontosubRecargo').text = value
                        print(f'MontosubRecargo : {value}')
                        col_index += 1

                    # TablaImpuestoAdicional = ET.SubElement(Item, 'TablaImpuestoAdicional')

                    ImpuestoAdicional_count = 2
                    while True :
                        ImpuestoAdicional_count -= 1
                        if ImpuestoAdicional_count < 0 : 
                            break
                        # ImpuestoAdicional =  ET.SubElement(TablaImpuestoAdicional, 'ImpuestoAdicional')                                   
                        # MontosubRecargo
                        value = str(sheet.cell(in_row, column=col_index).value)
                        if value == "#e" :
                            value = "" 

                        # ET.SubElement(ImpuestoAdicional, 'TipoImpuesto').text = value
                        print(f'TipoImpuesto : {value}')
                        col_index += 1
                    
                    OtraMonedaDetalle = ET.SubElement(Item, 'OtraMonedaDetalle')

                    # PrecioOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'PrecioOtraMoneda').text = value
                    print(f'PrecioOtraMoneda : {value}')
                    col_index += 1

                    # DescuentoOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'DescuentoOtraMoneda').text = value
                    print(f'DescuentoOtraMoneda : {value}')
                    col_index += 1

                    # RecargoOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'RecargoOtraMoneda').text = value
                    print(f'RecargoOtraMoneda : {value}')

                    col_index += 1
                    # MontoItemOtraMoneda
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(OtraMonedaDetalle, 'MontoItemOtraMoneda').text = value
                    print(f'MontoItemOtraMoneda : {value}')
                    col_index += 1
                    
                    # MontoItem
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = "" 

                    ET.SubElement(Item, 'MontoItem').text = value
                    print(f'MontoItem : {value}')
                    col_index += 1
        
        Subtotales = ET.SubElement(root, 'Subtotales')
        Subtotal = ET.SubElement(Subtotales, 'Subtotal')

        # Write NumeroSubTotal
        search_text = "NumeroSubTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'NumeroSubTotal').text = value
                print(f'Version : {cell_value}')
                break
        
        # Write DescripcionSubtotal
        search_text = "DescripcionSubtotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'DescripcionSubtotal').text = value
                print(f'DescripcionSubtotal : {cell_value}')
                break
        
        # Write Orden
        search_text = "Orden" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'Orden').text = value
                print(f'Orden : {cell_value}')
                break

        # Write SubTotalExento
        search_text = "SubTotalExento" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'SubTotalExento').text = value
                print(f'SubTotalExento : {cell_value}')
                break

        # Write MontoSubTotal
        search_text = "MontoSubTotal" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'MontoSubTotal').text = value
                print(f'MontoSubTotal : {cell_value}')
                break

        # Write Lineas
        search_text = "Lineas" 
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:
                value = str(sheet.cell(in_row, column=col).value)
                if value == "#e" :
                    value = ""             
                ET.SubElement(Subtotal, 'Lineas').text = value
                print(f'Lineas : {cell_value}')
                break

        Paginacion = ET.SubElement(root, 'Paginacion')
        Pagina_count = 2

        search_text = "PaginaNo[1]" 
        col_index : int
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column=col).value
            if cell_value == search_text:

                col_index = col
                    
                while True:
                    Pagina_count -= 1
                    if Pagina_count < 0:
                        break

                    Pagina = ET.SubElement(Paginacion, 'Pagina')

                    # PaginNo
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'PaginaNo').text = value
                    print(f'PaginaNo : {cell_value}')
                    print(f'PaginaNo col_index : {col_index}')
                    col_index += 1

                    # NoLineaDesde
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'NoLineaDesde').text = value
                    print(f'NoLineaDesde : {cell_value}')
                    col_index += 1

                    # NoLineaHasta
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'NoLineaHasta').text = value
                    print(f'NoLineaHasta : {value}')
                    col_index += 1
                    
                    # SubtotalMontoGravadoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravadoPagina').text = value
                    print(f'SubtotalMontoGravadoPagina : {value}')
                    col_index += 1
                                    
                    # SubtotalMontoGravado1Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravado1Pagina').text = value
                    print(f'SubtotalMontoGravado1Pagina : {value}')
                    col_index += 1
                                                    
                    # SubtotalMontoGravado2Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravado2Pagina').text = value
                    print(f'SubtotalMontoGravado2Pagina : {value}')
                    col_index += 1
                                                    
                    # SubtotalMontoGravado3Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalMontoGravado3Pagina').text = value
                    print(f'SubtotalMontoGravado3Pagina : {value}')
                    col_index += 1
                                                                    
                    # SubtotalExentoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    ET.SubElement(Pagina, 'SubtotalExentoPagina').text = value
                    print(f'SubtotalExentoPagina : {value}')
                    col_index += 1
                                                                
                    # SubtotalItbisPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbisPagina').text = value
                    print(f'SubtotalItbisPagina : {value}')
                    col_index += 1
                                                                        
                    # SubtotalItbis1Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbis1Pagina').text = value
                    print(f'SubtotalItbis1Pagina : {value}')
                    col_index += 1
                                                                                        
                    # SubtotalItbis2Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbis2Pagina').text = value
                    print(f'SubtotalItbis2Pagina : {value}')
                    col_index += 1
                                                                                            
                    # SubtotalItbis3Pagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalItbis3Pagina').text = value
                    print(f'SubtotalItbis3Pagina : {value}')
                    col_index += 1
                                                                                                            
                    # SubtotalImpuestoAdicionalPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""             
                    # ET.SubElement(Pagina, 'SubtotalImpuestoAdicionalPagina').text = value
                    print(f'SubtotalImpuestoAdicionalPagina : {value}')
                    col_index += 1

                    # SubtotalImpuestoAdicional = ET.SubElement(Pagina, 'SubtotalImpuestoAdicional')
                    
                    # SubtotalImpuestoSelectivoConsumoEspecificoPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalImpuestoSelectivoConsumoEspecificoPagina').text =value
                    print(f'SubtotalImpuestoSelectivoConsumoEspecificoPagina : {value}')
                    col_index += 1

                    # SubtotalOtrosImpuesto
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    # ET.SubElement(SubtotalImpuestoAdicional, 'SubtotalOtrosImpuesto').text =value
                    print(f'SubtotalOtrosImpuesto : {value}')
                    col_index += 1

                    # MontoSubtotalPagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    ET.SubElement(Pagina, 'MontoSubtotalPagina').text =value
                    print(f'MontoSubtotalPagina : {value}')
                    col_index += 1

                    # SubtotalMontoNoFacturablePagina
                    value = str(sheet.cell(in_row, column=col_index).value)
                    if value == "#e" :
                        value = ""     
                    # ET.SubElement(Pagina, 'SubtotalMontoNoFacturablePagina').text =value
                    print(f'SubtotalMontoNoFacturablePagina : {value}')
                    print(f'SubtotalMontoNoFacturablePagina _ index : {col_index}')
                    col_index += 1

        InformacionReferencia = ET.SubElement(root, 'InformacionReferencia')
        
        # NCFModificado
        col_index += 1
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'NCFModificado').text =value
        print(f'NCFModificado : {value}')
        col_index += 1
            
        # RNCOtroContribuyente
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'RNCOtroContribuyente').text =value
        print(f'RNCOtroContribuyente : {value}')
        col_index += 1
            
        # FechaNCFModificado
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'FechaNCFModificado').text =value
        print(f'FechaNCFModificado : {value}')
        col_index += 1
            
        # CodigoModificacion
        value = str(sheet.cell(in_row, column=col_index).value)
        if value == "#e" :
            value = ""     
        ET.SubElement(InformacionReferencia, 'CodigoModificacion').text =value
        print(f'CodigoModificacion : {value}')
        col_index += 1

        rough_string = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        pretty_xml_as_str = reparsed.toprettyxml(indent="  ", encoding='utf-8')

        self.invoice_name = f'{self.RNCEmisor}{self.ENCF}'
        path = os.path.join(os.path.dirname(__file__), f'data/{self.invoice_name}.xml')
        
        with open(path, 'wb') as f:
            f.write(pretty_xml_as_str)
        
        xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
        return xml_str
